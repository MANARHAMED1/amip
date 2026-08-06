import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parent / "kpi_specification.xlsx"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
section_font = Font(bold=True, size=11, color="2F5496")
future_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)


def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def style_row(ws, row, ncols, fill=None):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = wrap_align
        if fill:
            cell.fill = fill


# ─── All KPI data ────────────────────────────────────────────────────

KPI_DATA = [
    # ── Production ──
    {"id": "KPI-PRD-001", "name": "Estimated Machining Time per OF", "french": "Temps usinage prevu par OF",
     "section": "Production", "objective": "Know total planned machining time before production",
     "formula": "SUM(pg.temps_usinage_prevu) grouped by gamme_id",
     "tables": "phase_gamme, ordre_fabrication", "columns": "phase_gamme.temps_usinage_prevu, phase_gamme.gamme_id, ordre_fabrication.gamme_id",
     "output": "Integer", "unit": "Minutes", "interpretation": "Baseline for comparing actual. Used for scheduling",
     "typical": "30-300 min", "visual": "Card (per OF)", "refresh": "On OF creation", "phase": "Phase 1"},

    {"id": "KPI-PRD-002", "name": "Real Machining Time per OF", "french": "Temps usinage reel par OF",
     "section": "Production", "objective": "Measure actual machining time after execution",
     "formula": "SUM(ep.temps_usinage_reel) WHERE ordre_fabrication_id = OF AND statut = TERMINE",
     "tables": "execution_phase", "columns": "temps_usinage_reel, ordre_fabrication_id, statut",
     "output": "Integer", "unit": "Minutes", "interpretation": "Real time consumed. Compare with PRD-001",
     "typical": "35-350 min", "visual": "Card + comparison bar", "refresh": "After OF completion", "phase": "Phase 1"},

    {"id": "KPI-PRD-003", "name": "Estimated Setup Time per OF", "french": "Temps reglage prevu par OF",
     "section": "Production", "objective": "Track planned non-productive preparation time",
     "formula": "SUM(pg.temps_reglage_prevu) grouped by gamme",
     "tables": "phase_gamme", "columns": "phase_gamme.temps_reglage_prevu, phase_gamme.gamme_id",
     "output": "Integer", "unit": "Minutes", "interpretation": "Setup is non-value-added time. Lean priority to reduce",
     "typical": "5-60 min", "visual": "Card", "refresh": "On OF creation", "phase": "Phase 1"},

    {"id": "KPI-PRD-004", "name": "Real Setup Time per OF", "french": "Temps reglage reel par OF",
     "section": "Production", "objective": "Measure actual setup time",
     "formula": "SUM(ep.temps_reglage_reel) grouped by ordre_fabrication_id",
     "tables": "execution_phase", "columns": "temps_reglage_reel, ordre_fabrication_id",
     "output": "Integer", "unit": "Minutes", "interpretation": "Compare with PRD-003. Excess = tooling or process issues",
     "typical": "5-80 min", "visual": "Card + comparison bar", "refresh": "After OF completion", "phase": "Phase 1"},

    {"id": "KPI-PRD-005", "name": "Total Production Time per OF", "french": "Temps total production par OF",
     "section": "Production", "objective": "Total machine time consumed by a production order",
     "formula": "SUM(ep.temps_usinage_reel + ep.temps_reglage_reel) WHERE statut = TERMINE",
     "tables": "execution_phase", "columns": "temps_usinage_reel, temps_reglage_reel, ordre_fabrication_id, statut",
     "output": "Integer", "unit": "Minutes", "interpretation": "Core input for cost calculations and utilization",
     "typical": "40-400 min", "visual": "Card", "refresh": "After OF completion", "phase": "Phase 1"},

    {"id": "KPI-PRD-006", "name": "Production Duration (Calendar)", "french": "Duree de production (calendaire)",
     "section": "Production", "objective": "Measure elapsed calendar time of an OF",
     "formula": "OF.date_fin_reelle - OF.date_debut_reelle",
     "tables": "ordre_fabrication", "columns": "date_fin_reelle, date_debut_reelle",
     "output": "Integer", "unit": "Days", "interpretation": "Includes non-working time. Compare with planned for delivery",
     "typical": "3-30 days", "visual": "Bar chart (planned vs actual)", "refresh": "On OF completion", "phase": "Phase 1"},

    {"id": "KPI-PRD-007", "name": "Cycle Time per Phase", "french": "Temps cycle par phase",
     "section": "Production", "objective": "Time to produce one piece in a specific phase",
     "formula": "ep.temps_usinage_reel / ep.nb_pieces_produites WHERE nb_pieces_produites > 0",
     "tables": "execution_phase", "columns": "temps_usinage_reel, nb_pieces_produites",
     "output": "Decimal", "unit": "min/pc", "interpretation": "Core capacity metric. Increasing trend = degradation",
     "typical": "0.5-15 min/pc", "visual": "Line chart (trend)", "refresh": "After execution", "phase": "Phase 1"},

    {"id": "KPI-PRD-008", "name": "Production Efficiency", "french": "Efficacite de production",
     "section": "Production", "objective": "Compare estimated vs real machining time",
     "formula": "SUM(pg.temps_usinage_prevu) / SUM(ep.temps_usinage_reel) * 100 per OF",
     "tables": "phase_gamme, execution_phase", "columns": "phase_gamme.temps_usinage_prevu, execution_phase.temps_usinage_reel",
     "output": "Decimal", "unit": "%", "interpretation": "100% = as planned. >100% = faster. <100% = overruns",
     "typical": "85-120%", "visual": "Gauge + trend line", "refresh": "After OF completion", "phase": "Phase 1"},

    {"id": "KPI-PRD-009", "name": "Production Yield", "french": "Rendement de production",
     "section": "Production", "objective": "Proportion of good parts vs total produced",
     "formula": "(SUM(nb_pieces_produites) - SUM(nb_pieces_rebut)) / SUM(nb_pieces_produites) * 100",
     "tables": "execution_phase", "columns": "nb_pieces_produites, nb_pieces_rebut",
     "output": "Decimal", "unit": "%", "interpretation": "Direct quality measure. Target > 95%",
     "typical": "88-99%", "visual": "Gauge", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-PRD-010", "name": "Production Throughput", "french": "Debit de production",
     "section": "Production", "objective": "Output rate - good pieces per hour",
     "formula": "SUM(nb_produites - nb_rebut) / (SUM(temps_usinage + temps_reglage) / 60)",
     "tables": "execution_phase", "columns": "nb_pieces_produites, nb_pieces_rebut, temps_usinage_reel, temps_reglage_reel",
     "output": "Decimal", "unit": "pcs/h", "interpretation": "Key capacity indicator. Bottleneck detection",
     "typical": "10-100 pcs/h", "visual": "Bar chart (by machine)", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-PRD-011", "name": "Capacity Utilization", "french": "Taux d'utilisation capacite",
     "section": "Production", "objective": "How much available production time is used",
     "formula": "SUM(temps_usinage + temps_reglage) / available_time * 100 per machine per period",
     "tables": "execution_phase", "columns": "temps_usinage_reel, temps_reglage_reel, machine_id, date_debut",
     "output": "Decimal", "unit": "%", "interpretation": "Low = idle machines. Default: 8h/day, 22 days/month",
     "typical": "60-85%", "visual": "Heatmap (machine x day)", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-PRD-012", "name": "OF Delay", "french": "Retard OF",
     "section": "Production", "objective": "Measure delivery delay for production orders",
     "formula": "OF.date_fin_reelle - OF.date_fin_prevue (in days)",
     "tables": "ordre_fabrication", "columns": "date_fin_reelle, date_fin_prevue",
     "output": "Integer", "unit": "Days", "interpretation": "Positive = late, negative = early. Critical for delivery",
     "typical": "-5 to +10 days", "visual": "Bar chart (distribution)", "refresh": "Weekly", "phase": "Phase 1"},

    # ── Quality ──
    {"id": "KPI-QLT-001", "name": "Scrap Rate", "french": "Taux de rebut",
     "section": "Quality", "objective": "Proportion of defective parts",
     "formula": "SUM(nb_non_conformes) / SUM(nb_controles) * 100",
     "tables": "controle_qualite", "columns": "nb_non_conformes, nb_controles",
     "output": "Decimal", "unit": "%", "interpretation": "Primary quality metric. > 5% triggers investigation",
     "typical": "2-8%", "visual": "Gauge + line chart", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-QLT-002", "name": "First Pass Yield (FPY)", "french": "Rendement premiere passe",
     "section": "Quality", "objective": "Parts passing QC on first attempt",
     "formula": "(SUM(nb_controles) - SUM(nb_non_conformes)) / SUM(nb_controles) * 100",
     "tables": "controle_qualite", "columns": "nb_controles, nb_non_conformes",
     "output": "Decimal", "unit": "%", "interpretation": "Process capability without rework. Higher = lower cost",
     "typical": "90-98%", "visual": "Gauge", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-QLT-003", "name": "Conformity Rate", "french": "Taux de conformite",
     "section": "Quality", "objective": "Proportion of conforming parts",
     "formula": "SUM(nb_conformes) / SUM(nb_controles) * 100",
     "tables": "controle_qualite", "columns": "nb_conformes, nb_controles",
     "output": "Decimal", "unit": "%", "interpretation": "Tracked per period, machine, operator",
     "typical": "92-99%", "visual": "Gauge + trend", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-QLT-004", "name": "Dimensional Accuracy", "french": "Precision dimensionnelle",
     "section": "Quality", "objective": "How close parts are to target dimensions",
     "formula": "ABS(dimension_mesuree - dimension_cible) WHERE both NOT NULL",
     "tables": "controle_qualite", "columns": "dimension_mesuree, dimension_cible, tolerance_plus, tolerance_moins",
     "output": "Decimal", "unit": "mm", "interpretation": "Must stay within tolerance. Track avg and max deviation",
     "typical": "0.001-0.05 mm", "visual": "Box plot / scatter", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-QLT-005", "name": "Surface Roughness", "french": " Rugosite de surface",
     "section": "Quality", "objective": "Track surface finish quality",
     "formula": "AVG(rugosite_mesuree) WHERE NOT NULL grouped by machine, part, period",
     "tables": "controle_qualite", "columns": "rugosite_mesuree",
     "output": "Decimal", "unit": "Ra (um)", "interpretation": "Lower = smoother. Increasing trend = tool wear",
     "typical": "0.4-6.3 um", "visual": "Line chart (trend)", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-QLT-006", "name": "Defects by Machine", "french": "Defauts par machine",
     "section": "Quality", "objective": "Identify machines with highest defects",
     "formula": "SUM(cq.nb_non_conformes) GROUP BY ep.machine_id via execution_id",
     "tables": "controle_qualite, execution_phase", "columns": "nb_non_conformes, execution_id, machine_id",
     "output": "Table", "unit": "Count / %", "interpretation": "High = maintenance or calibration needed",
     "typical": "Varies", "visual": "Bar chart (top machines)", "refresh": "Weekly", "phase": "Phase 1"},

    {"id": "KPI-QLT-007", "name": "Defects by Tool", "french": "Defauts par outil",
     "section": "Quality", "objective": "Identify tools causing most defects",
     "formula": "SUM(cq.nb_non_conformes) GROUP BY ep.outil_id via execution_id",
     "tables": "controle_qualite, execution_phase", "columns": "nb_non_conformes, execution_id, outil_id",
     "output": "Table", "unit": "Count", "interpretation": "Worn tools produce more defects. Correlate with wear data",
     "typical": "Varies", "visual": "Bar chart", "refresh": "Weekly", "phase": "Phase 1"},

    {"id": "KPI-QLT-008", "name": "Defects by Material", "french": "Defauts par matiere",
     "section": "Quality", "objective": "Identify materials with highest defects",
     "formula": "SUM(cq.nb_non_conformes) GROUP BY p.matiere_id via piece_id",
     "tables": "controle_qualite, piece", "columns": "nb_non_conformes, piece_id, piece.matiere_id",
     "output": "Table", "unit": "Count", "interpretation": "Some materials harder to machine. Informs selection",
     "typical": "Varies", "visual": "Bar / pie chart", "refresh": "Monthly", "phase": "Phase 1"},

    {"id": "KPI-QLT-009", "name": "Defects by Operator", "french": "Defauts par operateur",
     "section": "Quality", "objective": "Identify operators associated with defects",
     "formula": "SUM(cq.nb_non_conformes) GROUP BY ep.operateur_id via execution_id",
     "tables": "controle_qualite, execution_phase", "columns": "nb_non_conformes, execution_id, operateur_id",
     "output": "Table", "unit": "Count", "interpretation": "May indicate training needs. Cross-ref with competence",
     "typical": "Varies", "visual": "Bar chart", "refresh": "Monthly", "phase": "Phase 1"},

    {"id": "KPI-QLT-010", "name": "Defects by Cause Category", "french": "Defauts par categorie de cause",
     "section": "Quality", "objective": "Classify defects by root cause",
     "formula": "SUM(cq.nb_non_conformes) GROUP BY cr.categorie via cause_rebut_id",
     "tables": "controle_qualite, cause_rebut", "columns": "nb_non_conformes, cause_rebut_id, cause_rebut.categorie",
     "output": "Table", "unit": "Count", "interpretation": "Highest category = priority for corrective action",
     "typical": "Varies", "visual": "Pie + bar chart", "refresh": "Monthly", "phase": "Phase 1"},

    {"id": "KPI-QLT-011", "name": "Defects by Production Order", "french": "Defauts par ordre de fabrication",
     "section": "Quality", "objective": "Identify OFs with highest scrap",
     "formula": "SUM(cq.nb_non_conformes) GROUP BY ep.ordre_fabrication_id",
     "tables": "controle_qualite, execution_phase", "columns": "nb_non_conformes, execution_id, ordre_fabrication_id",
     "output": "Table", "unit": "Count", "interpretation": "Flags problematic runs for root cause analysis",
     "typical": "Varies", "visual": "Bar chart (top 10)", "refresh": "Weekly", "phase": "Phase 1"},

    # ── OEE ──
    {"id": "KPI-OEE-001", "name": "Availability", "french": "Disponibilite",
     "section": "OEE", "objective": "Proportion of planned time machine is running",
     "formula": "running_time / (running_time + maintenance_downtime + idle_time) * 100",
     "tables": "execution_phase, maintenance, sensor_data", "columns": "temps_usinage_reel, temps_reglage_reel, maintenance.duree, sensor_data.statut_machine",
     "output": "Decimal", "unit": "%", "interpretation": "Target > 85%. Low = too much downtime",
     "typical": "80-95%", "visual": "Gauge", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-OEE-002", "name": "Performance", "french": "Performance",
     "section": "OEE", "objective": "Speed efficiency vs optimal",
     "formula": "SUM(pg.temps_usinage_prevu) / SUM(ep.temps_usinage_reel) * 100 WHERE TERMINE",
     "tables": "execution_phase, phase_gamme", "columns": "temps_usinage_prevu, temps_usinage_reel, phase_gamme_id",
     "output": "Decimal", "unit": "%", "interpretation": "> 100% = faster than planned. < 85% = slowdown",
     "typical": "85-110%", "visual": "Gauge", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-OEE-003", "name": "Quality Rate", "french": "Taux de qualite",
     "section": "OEE", "objective": "Proportion of good parts",
     "formula": "(SUM(nb_produites) - SUM(nb_rebut)) / SUM(nb_produites) * 100 per machine",
     "tables": "execution_phase", "columns": "nb_pieces_produites, nb_pieces_rebut, machine_id",
     "output": "Decimal", "unit": "%", "interpretation": "Target > 98%. Low quality pulls down OEE",
     "typical": "92-99%", "visual": "Gauge", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-OEE-004", "name": "OEE", "french": "OEE - Efficacite globale",
     "section": "OEE", "objective": "Single metric: Availability x Performance x Quality",
     "formula": "(OEE-001/100) * (OEE-002/100) * (OEE-003/100) * 100",
     "tables": "Derived from OEE-001, OEE-002, OEE-003", "columns": "(Indirect)",
     "output": "Decimal", "unit": "%", "interpretation": "World-class = 85%. <65% = needs improvement. 65-80% = good",
     "typical": "50-75%", "visual": "Gauge (primary) + trend", "refresh": "Daily", "phase": "Phase 1"},

    # ── Machine Performance ──
    {"id": "KPI-MCH-001", "name": "Machine Running Time", "french": "Temps de fonctionnement",
     "section": "Machine Performance", "objective": "Total productive time per machine per period",
     "formula": "SUM(temps_usinage_reel + temps_reglage_reel) WHERE machine_id = M AND TERMINE",
     "tables": "execution_phase", "columns": "temps_usinage_reel, temps_reglage_reel, machine_id, statut, date_debut",
     "output": "Integer", "unit": "Minutes", "interpretation": "Foundation for utilization and OEE",
     "typical": "Varies", "visual": "Card + bar chart", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-MCH-002", "name": "Machine Downtime (Sensor)", "french": "Arret machine (capteurs)",
     "section": "Machine Performance", "objective": "Time machine stopped/broken per sensor data",
     "formula": "COUNT(*) * 30 / 60 WHERE statut_machine IN (STOPPED, BROKEN)",
     "tables": "sensor_data", "columns": "statut_machine, machine_id, timestamp",
     "output": "Decimal", "unit": "Hours", "interpretation": "High = bottleneck. Compare across machines",
     "typical": "Varies", "visual": "Bar chart", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-MCH-003", "name": "Machine Maintenance Time", "french": "Temps maintenance machine",
     "section": "Machine Performance", "objective": "Total time under maintenance",
     "formula": "SUM(m.duree) WHERE machine_id = M AND date_debut IN [period]",
     "tables": "maintenance", "columns": "duree, machine_id, date_debut",
     "output": "Integer", "unit": "Minutes", "interpretation": "High = unreliable. Track trends for replacement",
     "typical": "Varies", "visual": "Card + bar chart", "refresh": "Monthly", "phase": "Phase 1"},

    {"id": "KPI-MCH-004", "name": "Machine Avg Cycle Time", "french": "Temps cycle moyen machine",
     "section": "Machine Performance", "objective": "Average time per piece on a specific machine",
     "formula": "SUM(temps_usinage_reel) / SUM(nb_pieces_produites) WHERE machine_id = M",
     "tables": "execution_phase", "columns": "temps_usinage_reel, nb_pieces_produites, machine_id",
     "output": "Decimal", "unit": "min/pc", "interpretation": "Increasing trend = degradation. Compare same-type machines",
     "typical": "Varies", "visual": "Line (trend) + bar (compare)", "refresh": "Weekly", "phase": "Phase 1"},

    {"id": "KPI-MCH-005", "name": "Machine Efficiency", "french": "Efficacite machine",
     "section": "Machine Performance", "objective": "Estimated vs actual time per machine",
     "formula": "SUM(pg.temps_usinage_prevu) / SUM(ep.temps_usinage_reel) * 100 WHERE machine_id = M",
     "tables": "execution_phase, phase_gamme", "columns": "temps_usinage_prevu, temps_usinage_reel, phase_gamme_id, machine_id",
     "output": "Decimal", "unit": "%", "interpretation": "> 100% = faster. < 85% = investigate",
     "typical": "85-120%", "visual": "Bar chart (by machine)", "refresh": "Monthly", "phase": "Phase 1"},

    # ── Tool Management ──
    {"id": "KPI-TOL-001", "name": "Tool Wear per Execution", "french": "Usure outil par execution",
     "section": "Tool Management", "objective": "Tool lifetime consumed in one execution",
     "formula": "eo.usure_fin - eo.usure_debut",
     "tables": "execution_outil", "columns": "usure_debut, usure_fin",
     "output": "Integer", "unit": "Minutes", "interpretation": "High = aggressive parameters or hard material",
     "typical": "Varies", "visual": "Line chart (per tool)", "refresh": "After execution", "phase": "Phase 1"},

    {"id": "KPI-TOL-002", "name": "Cumulative Tool Consumption", "french": "Consommation outil cumulee",
     "section": "Tool Management", "objective": "Total lifetime consumed across all uses",
     "formula": "SUM(usure_fin - usure_debut) GROUP BY outil_id",
     "tables": "execution_outil", "columns": "usure_debut, usure_fin, outil_id",
     "output": "Integer", "unit": "Minutes", "interpretation": "Compare with duree_vie_totale for remaining %",
     "typical": "Varies", "visual": "Bar chart (top consumers)", "refresh": "Weekly", "phase": "Phase 1"},

    {"id": "KPI-TOL-003", "name": "Tool Lifetime Percentage", "french": "Pourcentage vie outil",
     "section": "Tool Management", "objective": "How much of tool life is consumed",
     "formula": "usure_actuelle / duree_vie_totale * 100",
     "tables": "outil", "columns": "usure_actuelle, duree_vie_totale",
     "output": "Decimal", "unit": "%", "interpretation": "> 80% = replace soon. 100% = exhausted",
     "typical": "0-100%", "visual": "Histogram (distribution)", "refresh": "Weekly", "phase": "Phase 1"},

    {"id": "KPI-TOL-004", "name": "Tool Replacement Indicator", "french": "Indicateur remplacement outil",
     "section": "Tool Management", "objective": "Flag tools needing replacement",
     "formula": "CASE WHEN vie_restante <= vie_totale * 0.10 THEN REPLACE WHEN <= 0.25 THEN WARNING ELSE OK END",
     "tables": "outil", "columns": "duree_vie_restante, duree_vie_totale",
     "output": "String", "unit": "Category", "interpretation": "REPLACE = immediate. WARNING = schedule",
     "typical": "Categorical", "visual": "Pie chart (OK/WARN/REPLACE)", "refresh": "Weekly", "phase": "Phase 1"},

    {"id": "KPI-TOL-005", "name": "Avg Tool Lifetime by Type", "french": "Vie moyenne outil par type",
     "section": "Tool Management", "objective": "Benchmark tool durability by type",
     "formula": "AVG(duree_vie_totale) GROUP BY type_outil",
     "tables": "outil", "columns": "duree_vie_totale, type_outil",
     "output": "Decimal", "unit": "Minutes", "interpretation": "Low for a type = consider alternatives",
     "typical": "Varies", "visual": "Bar chart", "refresh": "Monthly", "phase": "Phase 1"},

    {"id": "KPI-TOL-006", "name": "Tool Utilization Rate", "french": "Taux utilisation outil",
     "section": "Tool Management", "objective": "How actively each tool type is used",
     "formula": "COUNT(DISTINCT exec_id per type) / COUNT(DISTINCT exec_id total) * 100",
     "tables": "execution_outil, outil", "columns": "execution_id, outil_id, outil.type_outil",
     "output": "Decimal", "unit": "%", "interpretation": "Low = excess inventory. High = bottleneck risk",
     "typical": "Varies", "visual": "Bar chart", "refresh": "Monthly", "phase": "Phase 1"},

    # ── Inventory ──
    {"id": "KPI-INV-001", "name": "Stock Level (Parts)", "french": "Niveau stock pieces",
     "section": "Inventory", "objective": "Finished parts in stock",
     "formula": "sp.quantite_stock",
     "tables": "stock_piece", "columns": "quantite_stock, piece_id",
     "output": "Integer", "unit": "Pieces", "interpretation": "Zero = stockout. Very high = excess capital",
     "typical": "0-200", "visual": "Table + conditional fmt", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-INV-002", "name": "Stock Level (Materials)", "french": "Niveau stock matieres",
     "section": "Inventory", "objective": "Raw material quantities",
     "formula": "sm.quantite_stock",
     "tables": "stock_matiere", "columns": "quantite_stock, matiere_id",
     "output": "Decimal", "unit": "kg", "interpretation": "Below threshold = reorder needed",
     "typical": "50-2000 kg", "visual": "Table + conditional fmt", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-INV-003", "name": "Stock Level (Tools)", "french": "Niveau stock outils",
     "section": "Inventory", "objective": "Tool inventory quantities",
     "formula": "so.quantite_stock",
     "tables": "stock_outil", "columns": "quantite_stock, outil_id",
     "output": "Integer", "unit": "Pieces", "interpretation": "Below threshold = urgent reorder",
     "typical": "1-15", "visual": "Table + conditional fmt", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-INV-004", "name": "Stock Status", "french": "Etat du stock",
     "section": "Inventory", "objective": "Classify inventory health",
     "formula": "CASE WHEN qty <= 0 THEN OUT_OF_STOCK WHEN qty <= seuil THEN LOW WHEN qty > seuil*3 THEN OVERSTOCK ELSE NORMAL END",
     "tables": "stock_matiere, stock_outil, stock_piece", "columns": "quantite_stock, seuil_alerte",
     "output": "String", "unit": "Category", "interpretation": "OUT_OF_STOCK = immediate. LOW = reorder. OVERSTOCK = reduce",
     "typical": "Categorical", "visual": "Donut chart", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-INV-005", "name": "Material Consumption (Est.)", "french": "Consommation matiere (est.)",
     "section": "Inventory", "objective": "Estimate material consumed from production",
     "formula": "SUM(p.poids * ep.nb_pieces_produites) GROUP BY p.matiere_id WHERE TERMINE",
     "tables": "piece, execution_phase, ordre_fabrication", "columns": "piece.poids, nb_pieces_produites, ordre_fabrication.piece_id",
     "output": "Decimal", "unit": "kg", "interpretation": "Estimate only. Does not include machining waste",
     "typical": "Varies", "visual": "Bar chart (by material)", "refresh": "Weekly", "phase": "Phase 1"},

    {"id": "KPI-INV-006", "name": "Reorder Indicator", "french": "Indicateur reapprovisionnement",
     "section": "Inventory", "objective": "Flag items needing reorder",
     "formula": "quantite_stock <= seuil_alerte",
     "tables": "stock_matiere, stock_outil", "columns": "quantite_stock, seuil_alerte",
     "output": "Boolean", "unit": "Boolean", "interpretation": "TRUE = needs reorder. Trigger for purchasing",
     "typical": "TRUE/FALSE", "visual": "Alert list / table", "refresh": "Daily", "phase": "Phase 1"},

    {"id": "KPI-INV-007", "name": "Stock Value (Parts)", "french": "Valeur stock pieces",
     "section": "Inventory", "objective": "Estimate finished parts inventory value",
     "formula": "SUM(sp.quantite_stock * p.prix_revient)",
     "tables": "stock_piece, piece", "columns": "stock_piece.quantite_stock, piece.prix_revient",
     "output": "Decimal", "unit": "EUR", "interpretation": "Financial indicator. High = capital tied up",
     "typical": "Varies", "visual": "Card + bar chart", "refresh": "Monthly", "phase": "Phase 1"},

    # ── Maintenance ──
    {"id": "KPI-MNT-001", "name": "Maintenance Count by Machine", "french": "Nombre maintenances par machine",
     "section": "Maintenance", "objective": "Maintenance events per machine",
     "formula": "COUNT(*) GROUP BY machine_id WHERE date_debut IN [period]",
     "tables": "maintenance", "columns": "machine_id, date_debut",
     "output": "Integer", "unit": "Count", "interpretation": "High = unreliable. Compare same age/type",
     "typical": "Varies", "visual": "Bar chart", "refresh": "Monthly", "phase": "Phase 1"},

    {"id": "KPI-MNT-002", "name": "Preventive Maintenance Ratio", "french": "Ratio maintenance preventive",
     "section": "Maintenance", "objective": "Proportion of planned maintenance",
     "formula": "COUNT(preventive types) / COUNT(*) * 100",
     "tables": "maintenance", "columns": "type_maintenance",
     "output": "Decimal", "unit": "%", "interpretation": "Target > 70%. Low = reactive culture",
     "typical": "50-80%", "visual": "Gauge", "refresh": "Monthly", "phase": "Phase 1"},

    {"id": "KPI-MNT-003", "name": "Corrective Maintenance Ratio", "french": "Ratio maintenance corrective",
     "section": "Maintenance", "objective": "Proportion of unplanned interventions",
     "formula": "COUNT(Corrective + Remplacement roulement) / COUNT(*) * 100",
     "tables": "maintenance", "columns": "type_maintenance",
     "output": "Decimal", "unit": "%", "interpretation": "Target < 30%. High = unreliable machines",
     "typical": "20-50%", "visual": "Gauge", "refresh": "Monthly", "phase": "Phase 1"},

    {"id": "KPI-MNT-004", "name": "Machine Downtime (Maint.)", "french": "Arret machine (maintenance)",
     "section": "Maintenance", "objective": "Total downtime from maintenance",
     "formula": "SUM(m.duree) WHERE machine_id = M AND date_debut IN [period]",
     "tables": "maintenance", "columns": "duree, machine_id, date_debut",
     "output": "Integer", "unit": "Minutes", "interpretation": "Directly reduces OEE availability",
     "typical": "Varies", "visual": "Stacked bar (by type)", "refresh": "Monthly", "phase": "Phase 1"},

    {"id": "KPI-MNT-005", "name": "MTBF", "french": "MTBF - Temps moyen entre pannes",
     "section": "Maintenance", "objective": "Operating time between failures",
     "formula": "SUM(temps_usinage + temps_reglage) / COUNT(corrective events) per machine",
     "tables": "execution_phase, maintenance", "columns": "temps_usinage_reel, temps_reglage_reel, type_maintenance, machine_id",
     "output": "Decimal", "unit": "Hours", "interpretation": "Higher = more reliable. Declining = aging",
     "typical": "200-2000 h", "visual": "Card + line chart", "refresh": "Monthly", "phase": "Phase 1"},

    {"id": "KPI-MNT-006", "name": "MTTR", "french": "MTTR - Temps moyen de reparation",
     "section": "Maintenance", "objective": "Average corrective repair duration",
     "formula": "AVG(m.duree) WHERE type IN (Corrective, Remplacement roulement) per machine",
     "tables": "maintenance", "columns": "duree, type_maintenance",
     "output": "Decimal", "unit": "Minutes", "interpretation": "Lower = faster repairs, better spares stock",
     "typical": "60-360 min", "visual": "Card + bar chart", "refresh": "Monthly", "phase": "Phase 1"},

    {"id": "KPI-MNT-007", "name": "Maintenance Cost per Machine", "french": "Cout maintenance par machine",
     "section": "Maintenance", "objective": "Total maintenance spend per machine",
     "formula": "SUM(m.cout) WHERE machine_id = M AND date_debut IN [period]",
     "tables": "maintenance", "columns": "cout, machine_id, date_debut",
     "output": "Decimal", "unit": "EUR", "interpretation": "High cost + old machine = replacement candidate",
     "typical": "Varies", "visual": "Bar + trend line", "refresh": "Monthly", "phase": "Phase 1"},

    {"id": "KPI-MNT-008", "name": "Maintenance Frequency", "french": "Frequence maintenance",
     "section": "Maintenance", "objective": "Events per machine per month",
     "formula": "COUNT(*) / number_of_months WHERE machine_id = M",
     "tables": "maintenance", "columns": "machine_id, date_debut",
     "output": "Decimal", "unit": "events/month", "interpretation": "High = unreliable. Sudden increase = impending failure",
     "typical": "Varies", "visual": "Line chart (trend)", "refresh": "Monthly", "phase": "Phase 1"},

    # ── Sensor ──
    {"id": "KPI-SNS-001", "name": "Temperature Status", "french": "Etat temperature",
     "section": "Sensor", "objective": "Monitor spindle temperature health",
     "formula": "CASE WHEN temperature > 55 THEN CRITICAL WHEN > 45 THEN WARNING ELSE NORMAL END",
     "tables": "sensor_data", "columns": "temperature",
     "output": "String", "unit": "Category (deg C)", "interpretation": "Normal: 20-40. Warn: >45. Crit: >55",
     "typical": "20-55 deg C", "visual": "Gauge + line with thresholds", "refresh": "Real-time", "phase": "Phase 1"},

    {"id": "KPI-SNS-002", "name": "Vibration Status", "french": "Etat vibration",
     "section": "Sensor", "objective": "Monitor machine vibration",
     "formula": "CASE WHEN vibration > 2.5 THEN CRITICAL WHEN > 1.5 THEN WARNING ELSE NORMAL END",
     "tables": "sensor_data", "columns": "vibration",
     "output": "String", "unit": "Category (mm/s)", "interpretation": "Normal: 0.1-1.0. Warn: >1.5. Crit: >2.5",
     "typical": "0.1-3.5 mm/s", "visual": "Gauge + line with thresholds", "refresh": "Real-time", "phase": "Phase 1"},

    {"id": "KPI-SNS-003", "name": "Spindle Load Status", "french": "Etat charge broche",
     "section": "Sensor", "objective": "Monitor machine load",
     "formula": "CASE WHEN charge_frappe > 90 THEN CRITICAL WHEN > 80 THEN WARNING ELSE NORMAL END",
     "tables": "sensor_data", "columns": "charge_frappe",
     "output": "String", "unit": "Category (%)", "interpretation": "Normal: 25-70%. Warn: >80. Crit: >90",
     "typical": "0-95%", "visual": "Gauge + line chart", "refresh": "Real-time", "phase": "Phase 1"},

    {"id": "KPI-SNS-004", "name": "Anomaly Composite Score", "french": "Score composite anomalie",
     "section": "Sensor", "objective": "Single multi-sensor health indicator",
     "formula": "(temp_flag*1) + (vib_flag*2) + (load_flag*1.5) + (rpm_flag*1.5). Each 0/1/2",
     "tables": "sensor_data, machine", "columns": "temperature, vibration, charge_frappe, rpm, machine.rpm_max",
     "output": "Decimal", "unit": "Index (0-7)", "interpretation": "Score >= 3 = MACHINE AT RISK",
     "typical": "0-7", "visual": "Gauge + heatmap", "refresh": "Real-time", "phase": "Phase 1"},

    # ── Cost ──
    {"id": "KPI-CST-001", "name": "Material Cost per OF", "french": "Cout matiere par OF",
     "section": "Cost", "objective": "Raw material cost for a production order",
     "formula": "p.poids * mat.prix_kg * of.quantite_produite",
     "tables": "ordre_fabrication, piece, matiere", "columns": "piece.poids, matiere.prix_kg, ordre_fabrication.quantite_produite, ordre_fabrication.piece_id",
     "output": "Decimal", "unit": "EUR", "interpretation": "Largest cost component typically",
     "typical": "Varies", "visual": "Card + bar chart", "refresh": "Per OF", "phase": "Phase 1"},

    {"id": "KPI-CST-002", "name": "Tool Cost per OF", "french": "Cout outil par OF",
     "section": "Cost", "objective": "Tool consumption cost",
     "formula": "SUM(duree_utilisation / duree_vie_totale * cout_remplacement) per OF",
     "tables": "execution_outil, outil, execution_phase", "columns": "duree_utilisation, duree_vie_totale, cout_remplacement, ordre_fabrication_id",
     "output": "Decimal", "unit": "EUR", "interpretation": "Fraction of tool lifetime x replacement cost",
     "typical": "Varies", "visual": "Card", "refresh": "Per OF", "phase": "Phase 1"},

    {"id": "KPI-CST-003", "name": "Machining Cost per OF", "french": "Cout usinage par OF",
     "section": "Cost", "objective": "Machine time cost",
     "formula": "SUM(temps_usinage + temps_reglage) / 60 * hourly_rate per OF",
     "tables": "execution_phase", "columns": "temps_usinage_reel, temps_reglage_reel, ordre_fabrication_id",
     "output": "Decimal", "unit": "EUR", "interpretation": "Hourly rate is external config (30-80 EUR/h)",
     "typical": "Varies", "visual": "Card", "refresh": "Per OF", "phase": "Phase 1"},

    {"id": "KPI-CST-004", "name": "Maintenance Cost per Machine", "french": "Cout maintenance par machine",
     "section": "Cost", "objective": "Total maintenance expenditure (ref: KPI-MNT-007)",
     "formula": "SUM(m.cout) WHERE machine_id = M AND date_debut IN [period]",
     "tables": "maintenance", "columns": "cout, machine_id, date_debut",
     "output": "Decimal", "unit": "EUR", "interpretation": "Part of total production cost",
     "typical": "Varies", "visual": "Bar + trend", "refresh": "Monthly", "phase": "Phase 1"},

    {"id": "KPI-CST-005", "name": "Cost per Part", "french": "Cout unitaire",
     "section": "Cost", "objective": "Unit cost of one good piece",
     "formula": "(Material + Tool + Machining + allocated Maintenance) / (qty_produced - qty_scrapped)",
     "tables": "Derived from CST-001 to CST-004, ordre_fabrication", "columns": "quantite_produite, quantite_rebut",
     "output": "Decimal", "unit": "EUR/pc", "interpretation": "Compare with prix_revient for margin analysis",
     "typical": "Varies", "visual": "Card + trend line", "refresh": "Per OF", "phase": "Phase 1"},

    # ── Future KPIs ──
    {"id": "KPI-FUT-001", "name": "Inventory Turnover Rate", "french": "Taux rotation stock",
     "section": "Future", "objective": "How often inventory turns over",
     "formula": "SUM(consumption_movements) / AVG(stock_level) per period",
     "tables": "mouvement_stock (NOT YET CREATED)", "columns": "type_mouvement, quantite, date_mouvement",
     "output": "Decimal", "unit": "turns/period", "interpretation": "Requires mouvement_stock table",
     "typical": "4-12/year", "visual": "Line chart", "refresh": "Monthly", "phase": "Phase 2"},

    {"id": "KPI-FUT-002", "name": "Stock Coverage (Days)", "french": "Couverture stock (jours)",
     "section": "Future", "objective": "How many days stock will last",
     "formula": "current_stock / average_daily_consumption",
     "tables": "mouvement_stock (NOT YET CREATED)", "columns": "Requires historical consumption rate",
     "output": "Decimal", "unit": "Days", "interpretation": "Requires mouvement_stock for precise consumption",
     "typical": "7-60 days", "visual": "Card", "refresh": "Weekly", "phase": "Phase 2"},

    {"id": "KPI-FUT-003", "name": "Power Consumption Cost", "french": "Cout consommation electrique",
     "section": "Future", "objective": "Electricity cost per machine",
     "formula": "SUM(puissance * interval_hours) * electricity_rate",
     "tables": "sensor_data + config (NOT YET CREATED)", "columns": "puissance + electricity rate config",
     "output": "Decimal", "unit": "EUR", "interpretation": "Requires electricity rate config table",
     "typical": "Varies", "visual": "Bar chart", "refresh": "Monthly", "phase": "Phase 2"},

    {"id": "KPI-FUT-004", "name": "Overall Factory OEE", "french": "OEE usine global",
     "section": "Future", "objective": "Plant-wide weighted OEE",
     "formula": "Weighted aggregation of per-machine OEE",
     "tables": "execution_phase + shift schedule (NOT YET CREATED)", "columns": "Requires planned_production_time per machine",
     "output": "Decimal", "unit": "%", "interpretation": "Needs shift schedule configuration",
     "typical": "50-75%", "visual": "Gauge", "refresh": "Daily", "phase": "Phase 2"},

    {"id": "KPI-FUT-005", "name": "Planned Maint. Compliance", "french": "Conformite maint. planifiee",
     "section": "Future", "objective": "On-time preventive maintenance rate",
     "formula": "COUNT(on-time events) / COUNT(planned events) * 100",
     "tables": "maintenance_plan (NOT YET CREATED)", "columns": "Requires planned dates",
     "output": "Decimal", "unit": "%", "interpretation": "Needs maintenance_plan table",
     "typical": "Target > 90%", "visual": "Gauge", "refresh": "Monthly", "phase": "Phase 2"},

    {"id": "KPI-FUT-006", "name": "Scrap Cost", "french": "Cout rebut",
     "section": "Future", "objective": "Financial cost of defective parts",
     "formula": "SUM(nb_rebut * (material_cost + machining_cost) per part)",
     "tables": "execution_phase + cost allocation (PARTIAL)", "columns": "nb_pieces_rebut + allocated costs",
     "output": "Decimal", "unit": "EUR", "interpretation": "Needs proper cost allocation logic",
     "typical": "Varies", "visual": "Card + bar chart", "refresh": "Monthly", "phase": "Phase 2"},

    {"id": "KPI-FUT-007", "name": "Tool Cost per Part", "french": "Cout outil par piece",
     "section": "Future", "objective": "Tool cost allocated to individual parts",
     "formula": "tool_cost_per_of / quantity_good_produced",
     "tables": "execution_outil + full cost chain", "columns": "Requires OF -> piece -> execution -> tool costing",
     "output": "Decimal", "unit": "EUR/pc", "interpretation": "Needs full cost chain implementation",
     "typical": "Varies", "visual": "Card", "refresh": "Per OF", "phase": "Phase 2"},
]


# ─── Build workbook ────────────────────────────────────────────────────

wb = Workbook()
wb.remove(wb.active)

# ── Sheet 1: Summary ──
ws = wb.create_sheet("Summary", index=0)
headers = ["KPI ID", "KPI Name", "French Name", "Section", "Unit", "Refresh", "Phase"]
ws.append(headers)
style_header(ws, 1, len(headers))

for i, kpi in enumerate(KPI_DATA, 2):
    ws.append([kpi["id"], kpi["name"], kpi["french"], kpi["section"],
               kpi["unit"], kpi["refresh"], kpi["phase"]])
    fill = future_fill if kpi["phase"] == "Phase 2" else None
    style_row(ws, i, len(headers), fill)

ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 35
ws.column_dimensions["C"].width = 35
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 20
ws.column_dimensions["G"].width = 10
ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = "A2"

# ── Sheet 2: Full Specification ──
ws2 = wb.create_sheet("Full Specification")
headers2 = ["KPI ID", "KPI Name", "French Name", "Section", "Business Objective",
            "Formula", "Source Tables", "Source Columns", "Output Type", "Unit",
            "Business Interpretation", "Typical Values", "Dashboard Visual",
            "Refresh Frequency", "Phase"]
ws2.append(headers2)
style_header(ws2, 1, len(headers2))

for i, kpi in enumerate(KPI_DATA, 2):
    ws2.append([
        kpi["id"], kpi["name"], kpi["french"], kpi["section"],
        kpi["objective"], kpi["formula"], kpi["tables"], kpi["columns"],
        kpi["output"], kpi["unit"], kpi["interpretation"], kpi["typical"],
        kpi["visual"], kpi["refresh"], kpi["phase"],
    ])
    fill = future_fill if kpi["phase"] == "Phase 2" else None
    style_row(ws2, i, len(headers2), fill)
    ws2.row_dimensions[i].height = 40

ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 30
ws2.column_dimensions["D"].width = 20
ws2.column_dimensions["E"].width = 45
ws2.column_dimensions["F"].width = 55
ws2.column_dimensions["G"].width = 35
ws2.column_dimensions["H"].width = 50
ws2.column_dimensions["I"].width = 14
ws2.column_dimensions["J"].width = 22
ws2.column_dimensions["K"].width = 50
ws2.column_dimensions["L"].width = 22
ws2.column_dimensions["M"].width = 30
ws2.column_dimensions["N"].width = 20
ws2.column_dimensions["O"].width = 10
ws2.auto_filter.ref = ws2.dimensions
ws2.freeze_panes = "A2"

# ── Sheet 3: By Section ──
sections = {}
for kpi in KPI_DATA:
    sec = kpi["section"]
    if sec not in sections:
        sections[sec] = []
    sections[sec].append(kpi)

ws3 = wb.create_sheet("By Section")
row = 1
for sec_name in ["Production", "Quality", "OEE", "Machine Performance",
                  "Tool Management", "Inventory", "Maintenance", "Sensor", "Cost", "Future"]:
    kpis = sections.get(sec_name, [])
    if not kpis:
        continue

    ws3.cell(row=row, column=1, value=sec_name.upper()).font = section_font
    for c in range(1, 5):
        ws3.cell(row=row, column=c).fill = section_fill
        ws3.cell(row=row, column=c).border = thin_border
    row += 1

    sub_headers = ["KPI ID", "Name", "Unit", "Formula"]
    for ci, h in enumerate(sub_headers, 1):
        cell = ws3.cell(row=row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for kpi in kpis:
        ws3.cell(row=row, column=1, value=kpi["id"]).border = thin_border
        ws3.cell(row=row, column=2, value=kpi["name"]).border = thin_border
        ws3.cell(row=row, column=3, value=kpi["unit"]).border = thin_border
        ws3.cell(row=row, column=4, value=kpi["formula"]).border = thin_border
        for c in range(1, 5):
            ws3.cell(row=row, column=c).alignment = wrap_align
        if kpi["phase"] == "Phase 2":
            for c in range(1, 5):
                ws3.cell(row=row, column=c).fill = future_fill
        row += 1
    row += 1

ws3.column_dimensions["A"].width = 16
ws3.column_dimensions["B"].width = 40
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 80

# ── Sheet 4: ML Feature Mapping ──
ws4 = wb.create_sheet("ML Feature Mapping")
ml_headers = ["Model", "Role", "Feature / Target", "Source Table", "Source Column", "Type"]
ws4.append(ml_headers)
style_header(ws4, 1, len(ml_headers))

ML_DATA = [
    ("Scrap Prediction", "Target", "nb_pieces_rebut > 0", "execution_phase", "nb_pieces_rebut", "Binary / Regression"),
    ("Scrap Prediction", "Feature", "Cutting speed", "execution_phase", "vitesse_coupe", "Continuous"),
    ("Scrap Prediction", "Feature", "Feed rate", "execution_phase", "avance", "Continuous"),
    ("Scrap Prediction", "Feature", "Depth of cut", "execution_phase", "profondeur_passe", "Continuous"),
    ("Scrap Prediction", "Feature", "Tool wear at start", "execution_outil", "usure_debut", "Integer"),
    ("Scrap Prediction", "Feature", "Tool wear at end", "execution_outil", "usure_fin", "Integer"),
    ("Scrap Prediction", "Feature", "Tool remaining life %", "outil", "duree_vie_restante / duree_vie_totale", "Continuous"),
    ("Scrap Prediction", "Feature", "Tool type", "outil", "type_outil", "Categorical"),
    ("Scrap Prediction", "Feature", "Machine type", "machine", "type", "Categorical"),
    ("Scrap Prediction", "Feature", "Operator competence", "operateur", "niveau_competence", "Categorical"),
    ("Scrap Prediction", "Feature", "Material type", "matiere", "type_matiere", "Categorical"),
    ("Scrap Prediction", "Feature", "Part weight", "piece", "poids", "Continuous"),
    ("Scrap Prediction", "Feature", "Avg temperature (100)", "sensor_data", "AVG(temperature) rolling", "Continuous"),
    ("Scrap Prediction", "Feature", "Avg vibration (100)", "sensor_data", "AVG(vibration) rolling", "Continuous"),

    ("Machining Time Estimation", "Target", "temps_usinage_reel", "execution_phase", "temps_usinage_reel", "Integer"),
    ("Machining Time Estimation", "Feature", "Planned machining time", "phase_gamme", "temps_usinage_prevu", "Integer"),
    ("Machining Time Estimation", "Feature", "Planned setup time", "phase_gamme", "temps_reglage_prevu", "Integer"),
    ("Machining Time Estimation", "Feature", "Machine type", "machine", "type", "Categorical"),
    ("Machining Time Estimation", "Feature", "Machine RPM max", "machine", "rpm_max", "Integer"),
    ("Machining Time Estimation", "Feature", "Tool type", "outil", "type_outil", "Categorical"),
    ("Machining Time Estimation", "Feature", "Tool diameter", "outil", "diametre", "Continuous"),
    ("Machining Time Estimation", "Feature", "Material density", "matiere", "densite", "Continuous"),
    ("Machining Time Estimation", "Feature", "Part weight", "piece", "poids", "Continuous"),
    ("Machining Time Estimation", "Feature", "Operator competence", "operateur", "niveau_competence", "Categorical"),

    ("Predictive Maintenance", "Target", "maintenance within N days", "maintenance", "date_debut", "Time-to-event"),
    ("Predictive Maintenance", "Feature", "Rolling avg temperature", "sensor_data", "AVG(temperature)", "Continuous"),
    ("Predictive Maintenance", "Feature", "Rolling std temperature", "sensor_data", "STDDEV(temperature)", "Continuous"),
    ("Predictive Maintenance", "Feature", "Rolling avg vibration", "sensor_data", "AVG(vibration)", "Continuous"),
    ("Predictive Maintenance", "Feature", "Time since last maint.", "maintenance", "MAX(date_debut) per machine", "Derived"),
    ("Predictive Maintenance", "Feature", "Machine age (days)", "machine", "date_installation", "Derived"),
    ("Predictive Maintenance", "Feature", "Cumulative running hours", "execution_phase", "SUM(temps_usinage+reglage)", "Derived"),
    ("Predictive Maintenance", "Feature", "Corrective maint. count", "maintenance", "COUNT WHERE type=Corrective", "Derived"),

    ("Machine Failure Prediction", "Target", "statut_machine = BROKEN within N readings", "sensor_data", "statut_machine", "Binary"),
    ("Machine Failure Prediction", "Feature", "Temperature trend slope", "sensor_data", "Linear slope of temperature", "Continuous"),
    ("Machine Failure Prediction", "Feature", "Vibration trend slope", "sensor_data", "Linear slope of vibration", "Continuous"),
    ("Machine Failure Prediction", "Feature", "Anomaly composite score", "sensor_data", "KPI-SNS-004", "Continuous"),
    ("Machine Failure Prediction", "Feature", "Time since corrective maint.", "maintenance", "Derived", "Continuous"),
    ("Machine Failure Prediction", "Feature", "Machine utilization rate", "execution_phase", "Derived", "Continuous"),
    ("Machine Failure Prediction", "Feature", "RPM stability (std)", "sensor_data", "STDDEV(rpm)", "Continuous"),

    ("Tool Wear Prediction", "Target", "duree_vie_restante or usure_fin", "outil / execution_outil", "duree_vie_restante, usure_fin", "Regression"),
    ("Tool Wear Prediction", "Feature", "Tool type", "outil", "type_outil", "Categorical"),
    ("Tool Wear Prediction", "Feature", "Tool diameter", "outil", "diametre", "Continuous"),
    ("Tool Wear Prediction", "Feature", "Tool material", "outil", "matiere_outil", "Categorical"),
    ("Tool Wear Prediction", "Feature", "Cumulative usage count", "execution_outil", "COUNT per outil_id", "Derived"),
    ("Tool Wear Prediction", "Feature", "Recent wear rate (last 5)", "execution_outil", "AVG(usure_fin-usure_debut)", "Derived"),
    ("Tool Wear Prediction", "Feature", "Cutting speed", "execution_phase", "vitesse_coupe", "Continuous"),
    ("Tool Wear Prediction", "Feature", "Material type", "piece -> matiere", "type_matiere", "Categorical"),

    ("Production Duration Prediction", "Target", "date_fin_reelle - date_debut_reelle", "ordre_fabrication", "date_fin_reelle, date_debut_reelle", "Integer (days)"),
    ("Production Duration Prediction", "Feature", "Quantity demanded", "ordre_fabrication", "quantite_demandee", "Integer"),
    ("Production Duration Prediction", "Feature", "Number of phases", "gamme_usinage", "nb_phases", "Integer"),
    ("Production Duration Prediction", "Feature", "Estimated total time", "gamme_usinage", "duree_totale_estimee", "Integer"),
    ("Production Duration Prediction", "Feature", "Material type", "matiere", "type_matiere", "Categorical"),
    ("Production Duration Prediction", "Feature", "Machine status", "machine", "statut", "Categorical"),
    ("Production Duration Prediction", "Feature", "Priority", "ordre_fabrication", "priorite", "Categorical"),

    ("Inventory Forecasting", "Target", "stock level at future dates", "stock_matiere / stock_piece", "quantite_stock", "Time-series"),
    ("Inventory Forecasting", "Feature", "Current stock level", "stock_matiere / stock_piece", "quantite_stock", "Continuous"),
    ("Inventory Forecasting", "Feature", "Daily consumption rate", "execution_phase + piece", "Derived from production", "Continuous"),
    ("Inventory Forecasting", "Feature", "Active OFs for this item", "ordre_fabrication", "COUNT WHERE statut IN (...)", "Integer"),
    ("Inventory Forecasting", "Feature", "Reorder threshold", "stock_matiere / stock_outil", "seuil_alerte", "Continuous"),
]

for i, (model, role, feature, table, column, ftype) in enumerate(ML_DATA, 2):
    ws4.append([model, role, feature, table, column, ftype])
    fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") if role == "Target" else None
    style_row(ws4, i, len(ml_headers), fill)

ws4.column_dimensions["A"].width = 32
ws4.column_dimensions["B"].width = 10
ws4.column_dimensions["C"].width = 35
ws4.column_dimensions["D"].width = 30
ws4.column_dimensions["E"].width = 40
ws4.column_dimensions["F"].width = 22
ws4.auto_filter.ref = ws4.dimensions
ws4.freeze_panes = "A2"

# ── Sheet 5: Dashboard Mapping ──
ws5 = wb.create_sheet("Dashboard Mapping")
dash_headers = ["KPI ID", "KPI Name", "Visualization", "Refresh", "Dashboard Section"]
ws5.append(dash_headers)
style_header(ws5, 1, len(dash_headers))

DASH_SECTIONS = {
    "KPI-PRD": "Production Overview",
    "KPI-QLT": "Quality Control",
    "KPI-OEE": "OEE Dashboard",
    "KPI-MCH": "Machine Monitor",
    "KPI-TOL": "Tool Management",
    "KPI-INV": "Inventory",
    "KPI-MNT": "Maintenance",
    "KPI-SNS": "Sensor Monitoring",
    "KPI-CST": "Financial",
    "KPI-FUT": "Future (Phase 2)",
}

for i, kpi in enumerate(KPI_DATA, 2):
    prefix = kpi["id"][:7]
    dash_section = DASH_SECTIONS.get(prefix, "Other")
    ws5.append([kpi["id"], kpi["name"], kpi["visual"], kpi["refresh"], dash_section])
    fill = future_fill if kpi["phase"] == "Phase 2" else None
    style_row(ws5, i, len(dash_headers), fill)

ws5.column_dimensions["A"].width = 16
ws5.column_dimensions["B"].width = 35
ws5.column_dimensions["C"].width = 40
ws5.column_dimensions["D"].width = 20
ws5.column_dimensions["E"].width = 25
ws5.auto_filter.ref = ws5.dimensions
ws5.freeze_panes = "A2"

# ── Save ──
wb.save(OUTPUT)
size_kb = OUTPUT.stat().st_size / 1024
print(f"Output: {OUTPUT} ({size_kb:.0f} KB)")
print(f"KPIs: {len(KPI_DATA)}")
print(f"Sheets: Summary, Full Specification, By Section, ML Feature Mapping, Dashboard Mapping")
