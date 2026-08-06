"""
AMIP - KPI Catalog Excel Generator
Generates docs/kpi_catalog.xlsx from embedded KPI data.
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles ──────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
BODY_FONT = Font(name="Calibri", size=10)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

KPI_COLUMNS = [
    "KPI ID", "Nom", "Module", "Question Metier",
    "Inputs Requis", "Information Retournee", "Formule",
    "Visualisation", "Interpretation",
    "Seuil Vert", "Seuil Orange", "Seuil Rouge", "Configurable",
    "Source Tables", "Source Colonnes", "Utilisation ML",
]

# ── KPI Data ────────────────────────────────────────────────────────────

EXEC_KPIS = [
    ("EXEC-001", "OEE Global", "Vue Executive", "Comment se porte l'efficacite globale de mon atelier ?", "Periode", "Moyenne ponderee de tous les OEE machines", "AVG(dwh.fact_execution.oee) * 100", "Carte + jauge", "Indicateur synthetique de la performance globale", ">= 75%", "60%-75%", "< 60%", "Oui", "dwh.fact_execution", "oee", "-"),
    ("EXEC-002", "Disponibilite Globale", "Vue Executive", "Quel pourcentage de temps mes machines sont-elles disponibles ?", "Periode", "Moyenne de la disponibilite", "AVG(taux_disponibilite) * 100", "Carte", "Capacite de production reelle vs temps total", ">= 85%", "70%-85%", "< 70%", "Oui", "dwh.fact_execution", "taux_disponibilite", "-"),
    ("EXEC-003", "Performance Globale", "Vue Executive", "Mes machines produisent-elles au rythme prevu ?", "Periode", "Moyenne de la performance", "AVG(taux_performance) * 100", "Carte", "Vitesse de production reelle vs theorique", ">= 90%", "75%-90%", "< 75%", "Oui", "dwh.fact_execution", "taux_performance", "-"),
    ("EXEC-004", "Qualite Globale", "Vue Executive", "Quel est le taux de conformite global ?", "Periode", "Taux de conformite moyen", "AVG(taux_qualite) * 100", "Carte", "Qualite globale de production", ">= 98%", "95%-98%", "< 95%", "Oui", "dwh.fact_execution", "taux_qualite", "-"),
    ("EXEC-005", "Production Totale", "Vue Executive", "Combien de pieces mon atelier a-t-il produites ?", "Periode", "Somme des pieces produites", "SUM(nb_pieces_produites)", "Carte", "Volume de production global", ">= Objectif", "80%-100%", "< 80%", "Oui", "dwh.fact_execution", "nb_pieces_produites", "-"),
    ("EXEC-006", "Rebut Total", "Vue Executive", "Combien de pieces ai-je perdues ?", "Periode", "Somme des rebuts", "SUM(nb_pieces_rebut)", "Carte", "Perte de production", "<= 2%", "2%-5%", "> 5%", "Oui", "dwh.fact_execution", "nb_pieces_rebut", "-"),
    ("EXEC-007", "Taux de Rebut Global", "Vue Executive", "Quel est le pourcentage de rebuts ?", "Periode", "Pourcentage de rebuts", "SUM(rebut)/SUM(produite)*100", "Carte + jauge", "Indicateur qualite global", "<= 2%", "2%-5%", "> 5%", "Oui", "dwh.fact_execution", "nb_pieces_rebut, nb_pieces_produites", "-"),
    ("EXEC-008", "OFs Actifs", "Vue Executive", "Combien d'OFs sont en production ?", "Periode", "Nombre d'OFs EN_COURS", "COUNT(OF WHERE statut='EN_COURS')", "Carte", "Charge de production", "Selon capacite", "-", "-", "Non", "dwh.dim_production_order", "statut", "-"),
    ("EXEC-009", "Machines Disponibles", "Vue Executive", "Combien de machines sont operationnelles ?", "-", "Nombre de machines RUNNING", "COUNT(machine WHERE statut='RUNNING')", "Carte (sur 12)", "Capacite operationnelle", ">= 10/12", "7-9/12", "< 7/12", "Oui", "dwh.dim_machine", "statut", "-"),
    ("EXEC-010", "Alertes Critiques", "Vue Executive", "Y a-t-il des situations urgentes ?", "-", "Nombre d'alertes critiques", "COUNT(alerts WHERE severity='Critique')", "Carte (rouge)", "Urgences a traiter", "0", "1-3", "> 3", "Oui", "Systeme d'alertes", "-", "-"),
    ("EXEC-011", "Production vs Plan", "Vue Executive", "Suis-je en retard sur mon plan ?", "Periode", "Produite vs demandee", "SUM(produite)/SUM(demandee)*100", "Barres comparatives", "Respect du planning", ">= 100%", "80%-100%", "< 80%", "Oui", "dwh.fact_production", "quantite_produite, quantite_demandee", "-"),
    ("EXEC-012", "OFs en Retard", "Vue Executive", "Combien d'OFs ont depasse la date prevue ?", "Periode", "Nombre d'OFs en retard", "COUNT(OF WHERE fin_reelle > fin_prevue)", "Carte", "Retards de production", "0", "1-5", "> 5", "Oui", "dwh.dim_production_order", "date_fin_prevue, date_fin_reelle", "-"),
    ("EXEC-013", "Cout Maintenance Total", "Vue Executive", "Combien coute la maintenance atelier ?", "Periode", "Somme des couts maintenance", "SUM(fact_maintenance.cout)", "Carte", "Budget maintenance consomme", "<= Budget", "100%-120%", "> 120%", "Oui", "dwh.fact_maintenance", "cout", "-"),
    ("EXEC-014", "OEE par Machine", "Vue Executive", "Quelles sont mes meilleures machines ?", "Periode", "Classement OEE machines", "AVG(oee) par machine", "Barres horizontales", "Identification des goulots", "-", "-", "-", "Non", "dwh.fact_execution, dwh.dim_machine", "oee, code", "-"),
    ("EXEC-015", "Tendance Production", "Vue Executive", "La production est-elle en hausse ?", "Periode", "Production quotidienne", "SUM(nb_pieces_produites) GROUP BY date", "Courbe", "Tendance", "-", "-", "-", "Non", "dwh.fact_execution, dwh.dim_date", "nb_pieces_produites", "-"),
]

MCH_KPIS = [
    ("MCH-001", "Statut Machine", "Machine", "La machine est-elle en marche ?", "Machine", "RUNNING/STOPPED/MAINTENANCE/BROKEN", "machine.statut", "Carte coloree", "Etat operationnel", "RUNNING", "STOPPED/MAINTENANCE", "BROKEN", "Non", "machine", "statut", "-"),
    ("MCH-002", "OF Actuel", "Machine", "Quel OF est en cours ?", "Machine", "Numero OF, piece, quantite", "SELECT numero_of FROM OF JOIN exec", "Carte", "Production en cours", "-", "-", "-", "Non", "ordre_fabrication, execution_phase", "numero_of", "-"),
    ("MCH-003", "Operateur Actuel", "Machine", "Qui pilote la machine ?", "Machine", "Nom, prenom, competence", "SELECT operateur WHERE EN_COURS", "Carte", "Affectation operateur", "Senior/Confirme", "-", "Junior", "Oui", "operateur, execution_phase", "nom, niveau_competence", "-"),
    ("MCH-004", "Outil Actuel", "Machine", "Quel outil est utilise ?", "Machine", "Code, type, usure", "SELECT outil WHERE EN_COURS", "Carte", "Outil en service", "Usure < 50%", "50%-80%", "> 80%", "Oui", "outil, execution_phase", "code, type_outil, usure_actuelle", "ML-05"),
    ("MCH-005", "Temps Usinage Prevu", "Machine", "Temps de production prevu ?", "Machine, Periode", "Somme temps prevu (min)", "SUM(phase_gamme.temps_usinage_prevu)", "Carte", "Capacite planifiee", "-", "-", "-", "Non", "phase_gamme", "temps_usinage_prevu", "ML-02"),
    ("MCH-006", "Temps Usinage Reel", "Machine", "Temps de production reel ?", "Machine, Periode", "Somme temps reel (min)", "SUM(execution_phase.temps_usinage_reel)", "Carte + comparaison", "Temps reel consomme", "<= 105% prevu", "105%-120%", "> 120%", "Oui", "execution_phase", "temps_usinage_reel", "ML-02"),
    ("MCH-007", "Temps Reglage Prevu", "Machine", "Temps de reglage prevu ?", "Machine, Periode", "Somme reglage prevu (min)", "SUM(phase_gamme.temps_reglage_prevu)", "Carte", "Temps non productif planifie", "-", "-", "-", "Non", "phase_gamme", "temps_reglage_prevu", "-"),
    ("MCH-008", "Temps Reglage Reel", "Machine", "Temps de reglage reel ?", "Machine, Periode", "Somme reglage reel (min)", "SUM(execution_phase.temps_reglage_reel)", "Carte + comparaison", "Temps non productif reel", "<= 105% prevu", "105%-120%", "> 120%", "Oui", "execution_phase", "temps_reglage_reel", "-"),
    ("MCH-009", "Disponibilite Machine", "Machine", "La machine est-elle disponible ?", "Machine, Periode", "Pourcentage disponibilite", "usinage_reel/(usinage+reglage)*100", "Carte + jauge", "Capacite a produire", ">= 85%", "70%-85%", "< 70%", "Oui", "execution_phase", "temps_usinage_reel, temps_reglage_reel", "-"),
    ("MCH-010", "Performance Machine", "Machine", "La machine produit-elle au rythme prevu ?", "Machine, Periode", "Pourcentage performance", "(prevu*produite)/reel*100", "Carte + jauge", "Efficacite de production", ">= 90%", "75%-90%", "< 75%", "Oui", "execution_phase, phase_gamme", "temps, nb_pieces", "-"),
    ("MCH-011", "Qualite Machine", "Machine", "La machine produit-elle conforme ?", "Machine, Periode", "Pourcentage qualite", "(produite-rebut)/produite*100", "Carte + jauge", "Qualite de production", ">= 98%", "95%-98%", "< 95%", "Oui", "execution_phase", "nb_pieces_produites, nb_pieces_rebut", "-"),
    ("MCH-012", "OEE Machine", "Machine", "Quel est l'OEE de cette machine ?", "Machine, Periode", "OEE en %", "Dispo*Perf*Qualite*100", "Carte + jauge", "Efficacite globale machine", ">= 75%", "60%-75%", "< 60%", "Oui", "execution_phase (calcule)", "-", "-"),
    ("MCH-013", "Temperature Actuelle", "Machine", "La temperature est-elle normale ?", "Machine", "Derniere lecture (C)", "DERNIERE(sensor_data.temperature)", "Carte + courbe", "Etat thermique", "< 60C", "60-80C", "> 80C", "Oui", "sensor_data", "temperature", "ML-04"),
    ("MCH-014", "Vibration Actuelle", "Machine", "La vibration est-elle excessive ?", "Machine", "Derniere lecture (mm/s)", "DERNIERE(sensor_data.vibration)", "Carte + courbe", "Etat mecanique", "< 2.5 mm/s", "2.5-4.5", "> 4.5", "Oui", "sensor_data", "vibration", "ML-04"),
    ("MCH-015", "RPM Actuel", "Machine", "La broche tourne-t-elle correctement ?", "Machine", "Derniere lecture RPM", "DERNIERE(sensor_data.rpm)", "Carte + jauge", "Vitesse de rotation", "Plage prevue", "Legere deviation", "Deviation majeure", "Oui", "sensor_data", "rpm", "-"),
    ("MCH-016", "Puissance Consommee", "Machine", "La consommation est-elle normale ?", "Machine", "Derniere lecture (kW)", "DERNIERE(sensor_data.puissance)", "Carte + courbe", "Consommation energetique", "Plage normale", "Legere augmentation", "Anormale", "Oui", "sensor_data", "puissance", "-"),
    ("MCH-017", "Score d'Anomalie", "Machine", "La machine presente-t-elle des anomalies ?", "Machine", "Score 0-100", "Calcule sur capteurs", "Carte + jauge", "Sante globale", "< 30", "30-60", "> 60", "Oui", "sensor_data (calcule)", "-", "ML-04"),
    ("MCH-018", "Nb Interventions", "Machine", "Combien de maintenances ?", "Machine, Periode", "Nombre d'interventions", "COUNT(maintenance WHERE machine=X)", "Carte", "Frequence maintenance", "Faible", "-", "Trop frequent", "Oui", "maintenance", "-", "-"),
    ("MCH-019", "Cout Maintenance", "Machine", "Combien coute la maintenance ?", "Machine, Periode", "Somme des couts", "SUM(cout WHERE machine=X)", "Carte + courbe", "Budget maintenance machine", "<= Budget", "100%-120%", "> 120%", "Oui", "maintenance", "cout", "-"),
    ("MCH-020", "MTBF", "Machine", "Temps moyen entre pannes ?", "Machine, Periode", "MTBF en heures", "Temps fonctionnement/Nb pannes", "Carte", "Fiabilite", ">= 200h", "100-200h", "< 100h", "Oui", "maintenance, sensor_data", "-", "ML-03"),
]

OF_KPIS = [
    ("OF-001", "Statut OF", "OF", "Mon OF est-il en cours ?", "Numero OF", "EN_ATTENTE/EN_COURS/TERMINE/ANNULE", "ordre_fabrication.statut", "Carte coloree", "Etat de l'OF", "TERMINE", "EN_COURS", "ANNULE", "Non", "ordre_fabrication", "statut", "-"),
    ("OF-002", "Quantite Demandee", "OF", "Combien de pieces a produire ?", "Numero OF", "Nombre demande", "quantite_demandee", "Carte", "Objectif", "-", "-", "-", "Non", "ordre_fabrication", "quantite_demandee", "-"),
    ("OF-003", "Quantite Produite", "OF", "Combien de pieces produites ?", "Numero OF", "Nombre produit", "quantite_produite", "Carte", "Realisation", ">= Demandee", "80%-100%", "< 80%", "Oui", "ordre_fabrication", "quantite_produite", "ML-06"),
    ("OF-004", "Pieces Bonnes", "OF", "Combien de bonnes pieces ?", "Numero OF", "Produite - Rebut", "quantite_produite - quantite_rebut", "Carte", "Production utile", ">= Demandee", "80%-100%", "< 80%", "Oui", "ordre_fabrication", "-", "-"),
    ("OF-005", "Pieces Rebutees", "OF", "Combien de pieces perdues ?", "Numero OF", "Nombre de rebuts", "quantite_rebut", "Carte", "Perte", "0", "1%-5%", "> 5%", "Oui", "ordre_fabrication", "quantite_rebut", "ML-01"),
    ("OF-006", "Taux Rebut OF", "OF", "Quel est le taux de perte ?", "Numero OF", "Pourcentage rebut", "rebut/produite*100", "Carte + jauge", "Qualite de l'OF", "<= 2%", "2%-5%", "> 5%", "Oui", "ordre_fabrication", "-", "ML-01"),
    ("OF-007", "Avancement", "OF", "Ou en est l'avancement ?", "Numero OF", "Pourcentage completion", "produite/demandee*100", "Jauge", "Progression", ">= 100%", "50%-100%", "< 50%", "Oui", "ordre_fabrication", "-", "-"),
    ("OF-008", "Duree Prevue", "OF", "Combien de temps dure l'OF ?", "Numero OF", "Nombre de jours", "fin_prevue - debut_prevue", "Carte", "Planning", "-", "-", "-", "Non", "ordre_fabrication", "dates", "ML-06"),
    ("OF-009", "Duree Reelle", "OF", "Duree reelle de l'OF ?", "Numero OF", "Nombre de jours reels", "fin_reelle - debut_reelle", "Carte + comparaison", "Reel vs prevu", "<= Prevu", "100%-120%", "> 120%", "Oui", "ordre_fabrication", "dates", "ML-06"),
    ("OF-010", "Retard", "OF", "L'OF est-il en retard ?", "Numero OF", "Jours de retard", "fin_reelle - fin_prevue", "Carte", "Retard", "<= 0", "1-3 jours", "> 3 jours", "Oui", "ordre_fabrication", "dates", "-"),
    ("OF-011", "Machine Utilisee", "OF", "Quelle machine est utilisee ?", "Numero OF", "Code(s) machine", "SELECT DISTINCT machine.code", "Tableau", "Affectation", "-", "-", "-", "Non", "machine, execution_phase", "code", "-"),
    ("OF-012", "Operateur(s)", "OF", "Qui travaille sur l'OF ?", "Numero OF", "Liste operateurs", "SELECT DISTINCT operateur.nom", "Tableau", "Affectation", "-", "-", "-", "Non", "operateur, execution_phase", "nom", "-"),
    ("OF-013", "Efficacite OF", "OF", "L'OF a-t-il ete efficace ?", "Numero OF", "Efficacite en %", "produite/demandee * prevu/reel", "Carte + jauge", "Efficacite composite", ">= 90%", "75%-90%", "< 75%", "Oui", "ordre_fabrication, execution_phase", "-", "-"),
    ("OF-014", "Priorite", "OF", "Quelle est la priorite ?", "Numero OF", "HAUTE/NORMALE/BASSE", "ordre_fabrication.priorite", "Badge colore", "Priorite", "NORMALE/BASSE", "-", "HAUTE (si retard)", "Non", "ordre_fabrication", "priorite", "-"),
    ("OF-015", "Phases par Ordre", "OF", "Combien de phases ?", "Numero OF", "Nombre phases", "COUNT(execution_phase)", "Frise ou tableau", "Complexite", "-", "-", "-", "Non", "execution_phase, phase_gamme", "-", "-"),
]

QUA_KPIS = [
    ("QUA-001", "Nb Inspections", "Qualite", "Combien de controles ?", "Periode, (Piece/Machine/OF)", "Nombre de controles", "COUNT(controle_qualite)", "Carte", "Activite de controle", "-", "-", "-", "Non", "controle_qualite", "-", "-"),
    ("QUA-002", "Pieces Conformes", "Qualite", "Combien de pieces conformes ?", "Periode", "Nombre conformes", "SUM(nb_conformes)", "Carte", "Production conforme", "-", "-", "-", "Non", "controle_qualite", "nb_conformes", "-"),
    ("QUA-003", "Pieces Non Conformes", "Qualite", "Combien de non conformes ?", "Periode", "Nombre non conformes", "SUM(nb_non_conformes)", "Carte", "Perte qualite", "0", "-", "> Seuil", "Oui", "controle_qualite", "nb_non_conformes", "ML-01"),
    ("QUA-004", "Taux Conformite", "Qualite", "Quel % de conformes ?", "Periode", "Pourcentage conformite", "SUM(conformes)/SUM(controles)*100", "Carte + jauge", "Qualite globale", ">= 98%", "95%-98%", "< 95%", "Oui", "controle_qualite", "-", "-"),
    ("QUA-005", "Taux Rebut", "Qualite", "Quel % de rebuts ?", "Periode", "Pourcentage rebut", "SUM(non_conformes)/SUM(controles)*100", "Carte + jauge", "Perte qualite", "<= 2%", "2%-5%", "> 5%", "Oui", "controle_qualite", "-", "ML-01"),
    ("QUA-006", "FPY", "Qualite", "Rendement au 1er passage ?", "Periode", "Pourcentage FPY", "Conformes au 1er / Total * 100", "Carte + jauge", "Qualite processus", ">= 95%", "90%-95%", "< 90%", "Oui", "controle_qualite", "-", "-"),
    ("QUA-007", "Defauts par Categorie", "Qualite", "Grandes causes de defauts ?", "Periode", "Repartition 6 categories", "COUNT GROUP BY categorie", "Camembert", "Pareto", "-", "-", "-", "Non", "controle_qualite, cause_rebut", "categorie", "-"),
    ("QUA-008", "Top 5 Causes", "Qualite", "Causes les plus frequentes ?", "Periode", "Top 5 descriptions", "COUNT GROUP BY description LIMIT 5", "Barres (Pareto)", "Priorisation actions", "-", "-", "-", "Non", "controle_qualite, cause_rebut", "description", "-"),
    ("QUA-009", "Evolution Taux Rebut", "Qualite", "La qualite s'ameliore-t-elle ?", "Periode", "Taux rebut par jour", "SUM(non_conformes)/SUM(controles) GROUP BY date", "Courbe", "Tendance qualite", "Baisse", "Stable", "Hausse", "Oui", "controle_qualite", "date_controle", "-"),
    ("QUA-010", "Qualite par Machine", "Qualite", "Quelle machine rebute le plus ?", "Periode", "Taux rebut par machine", "SUM(non_conformes)/SUM(controles) GROUP BY machine", "Barres horizontales", "Performance machine", "-", "-", "-", "Non", "controle_qualite, execution_phase, machine", "code", "-"),
    ("QUA-011", "Qualite par Operateur", "Qualite", "Meilleur operateur en qualite ?", "Periode", "Taux rebut par operateur", "SUM(non_conformes)/SUM(controles) GROUP BY operateur", "Barres horizontales", "Performance operateur", "-", "-", "-", "Non", "controle_qualite, execution_phase, operateur", "nom", "-"),
    ("QUA-012", "Qualite par Outil", "Qualite", "Quel outil cause des defauts ?", "Periode", "Taux rebut par outil", "SUM(non_conformes)/SUM(controles) GROUP BY outil", "Barres horizontales", "Impact outil", "-", "-", "-", "Non", "controle_qualite, execution_phase, outil", "code", "-"),
    ("QUA-013", "Qualite par Matiere", "Qualite", "Quelle mattere pose problemes ?", "Periode", "Taux rebut par matiere", "SUM(non_conformes)/SUM(controles) GROUP BY type_matiere", "Barres", "Impact matiere", "-", "-", "-", "Non", "controle_qualite, piece, matiere", "type_matiere", "-"),
    ("QUA-014", "Ecart Dimensionnel", "Qualite", "Mes pieces sont-elles dans tolerances ?", "Periode, Piece", "Ecart moyen", "AVG(dimension_mesuree - dimension_cible)", "Carte", "Precision processus", "Dans tolerance", "Proche limites", "Hors tolerance", "Oui", "controle_qualite", "dimension_mesuree, dimension_cible", "-"),
    ("QUA-015", "Rugosite Moyenne", "Qualite", "Finition de surface acceptable ?", "Periode, Piece", "Rugosite moyenne Ra", "AVG(rugosite_mesuree)", "Carte", "Qualite surface", "Selon spec", "Proche limite", "Hors spec", "Oui", "controle_qualite", "rugosite_mesuree", "-"),
    ("QUA-016", "Qualite par Piece", "Qualite", "Quelle piece a le plus de defauts ?", "Periode", "Taux rebut par reference", "SUM(non_conformes)/SUM(controles) GROUP BY reference", "Barres horizontales", "Pieces a problemes", "-", "-", "-", "Non", "controle_qualite, piece", "reference", "-"),
    ("QUA-017", "Defauts par Famille", "Qualite", "Quelle famille pose problemes ?", "Periode", "Taux rebut par famille", "SUM(non_conformes)/SUM(controles) GROUP BY famille", "Barres", "Tendance famille", "-", "-", "-", "Non", "controle_qualite, piece", "famille", "-"),
    ("QUA-018", "Ratio Conforme/NonConforme", "Qualite", "Quel est le ratio ?", "Periode", "Ratio", "SUM(conformes)/SUM(non_conformes)", "Camembert", "Equilibre qualite", ">= 20:1", "10:1-20:1", "< 10:1", "Oui", "controle_qualite", "-", "-"),
    ("QUA-019", "Nb Controles par Inspection", "Qualite", "Combien de mesures par controle ?", "Periode", "Moyenne nb_controles", "AVG(nb_controles)", "Carte", "Profondeur controle", "-", "-", "-", "Non", "controle_qualite", "nb_controles", "-"),
    ("QUA-020", "Derniere Inspection", "Qualite", "Quand dernier controle ?", "Piece", "Date derniere inspection", "MAX(date_controle)", "Carte", "Fraicheur controle", "< 7 jours", "7-30 jours", "> 30 jours", "Oui", "controle_qualite", "date_controle", "-"),
]

INV_KPIS = [
    ("INV-001", "Stock Matiere", "Inventaire", "Combien de matiere en stock ?", "Matiere", "Quantite en kg", "stock_matiere.quantite_stock", "Carte + jauge", "Niveau stock", "> seuil*1.5", "seuil-1.5*seuil", "<= seuil", "Oui", "stock_matiere", "quantite_stock, seuil_alerte", "ML-07"),
    ("INV-002", "Statut Stock Matiere", "Inventaire", "Mon stock est-il critique ?", "Matiere", "Critique/Bas/Normal/Surstock", "Logique seuils", "Badge colore", "Etat stock", "Normal", "Bas/Surstock", "Critique", "Oui", "stock_matiere", "-", "ML-07"),
    ("INV-003", "Seuil Alert Matiere", "Inventaire", "Quel est le seuil minimum ?", "Matiere", "Seuil d'alerte", "stock_matiere.seuil_alerte", "Carte", "Limite basse", "-", "-", "-", "Oui", "stock_matiere", "seuil_alerte", "-"),
    ("INV-004", "Valeur Stock Matiere", "Inventaire", "Combien vaut mon stock ?", "Matiere", "Valeur en EUR", "quantite_stock * prix_kg", "Carte", "Investissement", "-", "-", "-", "Non", "stock_matiere, matiere", "quantite_stock, prix_kg", "-"),
    ("INV-005", "Stock Outil", "Inventaire", "Combien d'outils en stock ?", "Outil", "Nombre unites", "stock_outil.quantite_stock", "Carte + jauge", "Disponibilite", "> seuil", "= seuil", "< seuil", "Oui", "stock_outil", "quantite_stock, seuil_alerte", "-"),
    ("INV-006", "Statut Stock Outil", "Inventaire", "Stock outils suffisant ?", "Outil", "Critique/Bas/Normal/Surstock", "Logique seuils", "Badge colore", "Etat stock", "Normal", "Bas/Surstock", "Critique", "Oui", "stock_outil", "-", "-"),
    ("INV-007", "Stock Pieces Finies", "Inventaire", "Combien de pieces finies ?", "Piece", "Nombre unites", "stock_piece.quantite_stock", "Carte", "Stock fini", "> 0", "= 0", "-", "Oui", "stock_piece", "quantite_stock", "-"),
    ("INV-008", "Statut Stock Pieces", "Inventaire", "Stock pieces suffisant ?", "Piece", "Critique/Bas/Normal/Surstock", "Logique conso moyenne", "Badge colore", "Etat stock", "Normal", "Bas/Surstock", "Critique", "Oui", "stock_piece", "-", "-"),
    ("INV-009", "Jours Restants Matiere", "Inventaire", "Combien de jours de stock ?", "Matiere", "Nombre de jours", "stock / conso_moyenne_jour", "Carte", "Duree avant rupture", "> 30 jours", "15-30 jours", "< 15 jours", "Oui", "stock_matiere, execution_phase", "-", "ML-07"),
    ("INV-010", "Valeur Totale Inventaire", "Inventaire", "Valeur totale ?", "-", "Valeur EUR totale", "SUM(stocks * prix)", "Carte + camembert", "Investissement total", "-", "-", "-", "Non", "stock_matiere, stock_outil, stock_piece, matiere, outil, piece", "-", "-"),
    ("INV-011", "Articles Critiques", "Inventaire", "Combien d'articles critiques ?", "-", "Nombre critiques", "COUNT(WHERE stock<=seuil)", "Carte (rouge)", "Urgences", "0", "1-3", "> 3", "Oui", "stock_matiere, stock_outil", "-", "ML-07"),
    ("INV-012", "Emplacement Stock", "Inventaire", "Ou se trouve le stock ?", "Article", "Code emplacement", "stock_*.emplacement", "Carte", "Localisation", "-", "-", "-", "Non", "stock_*", "emplacement", "-"),
    ("INV-013", "Derniere MAJ Stock", "Inventaire", "Le stock est-il a jour ?", "Article", "Date derniere MAJ", "stock_*.date_derniere_maj", "Carte", "Fraicheur", "< 24h", "24h-7j", "> 7j", "Oui", "stock_*", "date_derniere_maj", "-"),
    ("INV-014", "Cout Matiere/Piece", "Inventaire", "Cout matiere par piece ?", "Piece", "Cout EUR", "poids * prix_kg", "Carte", "Cout unitaire", "-", "-", "-", "Non", "piece, matiere", "poids, prix_kg", "-"),
    ("INV-015", "Rotation Stock", "Inventaire", "Le stock tourne-t-il ?", "Article, Periode", "Nb rotations", "Conso_totale / Stock_moyen", "Carte", "Efficacite gestion", ">= 6/an", "3-6/an", "< 3/an", "Oui", "stock_matiere, execution_phase", "-", "-"),
]

TL_KPIS = [
    ("TL-001", "Pourcentage Usure", "Outil", "A quel point l'outil est-il use ?", "Outil", "Pourcentage usure", "usure_actuelle/duree_vie_totale*100", "Jauge", "Usure", "< 50%", "50%-80%", "> 80%", "Oui", "outil", "usure_actuelle, duree_vie_totale", "ML-05"),
    ("TL-002", "Duree Vie Restante", "Outil", "Combien de temps reste-t-il ?", "Outil", "Minutes restantes", "outil.duree_vie_restante", "Carte + jauge", "Vie restante", "> 50%", "20%-50%", "< 20%", "Oui", "outil", "duree_vie_restante, duree_vie_totale", "ML-05"),
    ("TL-003", "Nb Executions", "Outil", "Combien de fois utilise ?", "Outil", "Nombre executions", "COUNT(execution_outil)", "Carte", "Utilisation", "-", "-", "-", "Non", "execution_outil", "-", "-"),
    ("TL-004", "Cout/Execution", "Outil", "Cout par utilisation ?", "Outil", "Cout EUR", "cout_achat / nb_executions", "Carte", "Cout unitaire", "-", "-", "-", "Non", "outil, execution_outil", "cout_achat", "-"),
    ("TL-005", "Indicateur Remplacement", "Outil", "Dois-je remplacer ?", "Outil", "OK/WARNING/CRITICAL", "Usure > 80% -> CRITICAL", "Badge colore", "Necessite remplacement", "OK", "WARNING", "CRITICAL", "Oui", "outil", "usure_actuelle, duree_vie_totale", "ML-05"),
    ("TL-006", "Machine Actuelle", "Outil", "Sur quelle machine ?", "Outil", "Code machine", "SELECT machine.code WHERE EN_COURS", "Carte", "Affectation", "-", "-", "-", "Non", "machine, execution_phase, execution_outil", "code", "-"),
    ("TL-007", "Usure par Execution", "Outil", "Usure a chaque utilisation ?", "Outil", "Min/execution", "AVG(usure_fin - usure_debut)", "Courbe", "Taux usure", "Faible", "-", "Eleve", "Oui", "execution_outil", "usure_debut, usure_fin", "ML-05"),
    ("TL-008", "Cout Achat", "Outil", "Combien coute l'outil ?", "Outil", "Cout EUR", "outil.cout_achat", "Carte", "Investissement", "-", "-", "-", "Non", "outil", "cout_achat", "-"),
    ("TL-009", "Cout Remplacement", "Outil", "Cout de remplacement ?", "Outil", "Cout EUR", "outil.cout_remplacement", "Carte", "Cout remplacement", "-", "-", "-", "Non", "outil", "cout_remplacement", "-"),
    ("TL-010", "Disponibilite", "Outil", "L'outil est-il disponible ?", "Outil", "Disponible/Indisponible", "outil.disponible", "Badge", "Disponibilite", "Disponible", "-", "Indisponible", "Non", "outil", "disponible", "-"),
    ("TL-011", "Type Outil", "Outil", "Quel type d'outil ?", "Outil", "Type (Foret, Fraise...)", "outil.type_outil", "Carte", "Categorie", "-", "-", "-", "Non", "outil", "type_outil", "-"),
    ("TL-012", "Diametre", "Outil", "Quel diametre ?", "Outil", "Diametre mm", "outil.diametre", "Carte", "Spec", "-", "-", "-", "Non", "outil", "diametre", "-"),
    ("TL-013", "Derniere Utilisation", "Outil", "Quand utilise la derniere fois ?", "Outil", "Date", "MAX(execution_outil.created_at)", "Carte", "Fraicheur", "< 30 jours", "30-90 jours", "> 90 jours", "Oui", "execution_outil", "created_at", "-"),
    ("TL-014", "Historique Usures", "Outil", "Comment l'usure evolue-t-elle ?", "Outil", "Courbe usure", "usure_fin GROUP BY date", "Courbe", "Tendance usure", "Lineaire", "Accelere", "Extrapolation critique", "Oui", "execution_outil", "usure_fin", "ML-05"),
    ("TL-015", "Cout Cumule", "Outil", "Cout total d'utilisation ?", "Outil", "Cout cumule", "cout_achat + remplacements*cout_remplacement", "Barres cumulees", "Cout possession", "-", "-", "-", "Non", "outil", "-", "-"),
]

MNT_KPIS = [
    ("MNT-001", "MTBF", "Maintenance", "Temps moyen entre pannes ?", "Machine, Periode", "MTBF en heures", "Temps_fonctionnement/Nb_pannes", "Carte", "Fiabilite", ">= 200h", "100-200h", "< 100h", "Oui", "maintenance, sensor_data", "-", "ML-03"),
    ("MNT-002", "MTTR", "Maintenance", "Temps moyen de reparation ?", "Machine, Periode", "MTTR en heures", "Somme_durees_reparation/Nb_reparations", "Carte", "Reparabilite", "<= 2h", "2-4h", "> 4h", "Oui", "maintenance", "duree", "-"),
    ("MNT-003", "Disponibilite", "Maintenance", "La machine est-elle disponible ?", "Machine, Periode", "Pourcentage", "MTBF/(MTBF+MTTR)*100", "Carte + jauge", "Capacite operationnelle", ">= 95%", "90%-95%", "< 90%", "Oui", "MNT-001, MNT-002", "-", "-"),
    ("MNT-004", "Nb Interventions", "Maintenance", "Combien d'interventions ?", "Machine, Periode", "Nombre", "COUNT(maintenance)", "Carte", "Activite", "-", "-", "-", "Non", "maintenance", "-", "-"),
    ("MNT-005", "Cout Total", "Maintenance", "Cout total maintenance ?", "Machine, Periode", "Somme couts", "SUM(cout)", "Carte", "Budget consomme", "<= Budget", "100%-120%", "> 120%", "Oui", "maintenance", "cout", "ML-03"),
    ("MNT-006", "Cout Moyen/Intervention", "Maintenance", "Cout moyen ?", "Machine, Periode", "Cout moyen", "SUM(cout)/COUNT(maintenance)", "Carte", "Cout unitaire", "-", "-", "-", "Non", "maintenance", "cout", "-"),
    ("MNT-007", "Duree Totale Arret", "Maintenance", "Temps total arret ?", "Machine, Periode", "Heures totales", "SUM(duree)/60", "Carte", "Temps perdu", "<= 5%", "5%-10%", "> 10%", "Oui", "maintenance", "duree", "-"),
    ("MNT-008", "Ratio Preventif/Correctif", "Maintenance", "Ratio preventif vs correctif ?", "Machine, Periode", "Ratio", "COUNT(Preventive)/COUNT(Corrective)", "Camembert", "Strategie", ">= 3:1", "1:1-3:1", "< 1:1", "Oui", "maintenance", "type_maintenance", "-"),
    ("MNT-009", "Repartition par Type", "Maintenance", "Types de maintenance dominants ?", "Machine, Periode", "Nombre par type", "COUNT GROUP BY type_maintenance", "Camembert", "Profil maintenance", "-", "-", "-", "Non", "maintenance", "type_maintenance", "-"),
    ("MNT-010", "Cout par Type", "Maintenance", "Quel type coute le plus ?", "Machine, Periode", "Cout par type", "SUM(cout) GROUP BY type", "Barres", "Distribution couts", "-", "-", "-", "Non", "maintenance", "cout, type_maintenance", "-"),
    ("MNT-011", "Cout Cumule", "Maintenance", "Budget consomme ?", "Machine, Periode", "Cout cumule mensuel", "SUM(cout) cumule par mois", "Courbe", "Consommation budget", "-", "-", "-", "Non", "maintenance", "cout", "ML-03"),
    ("MNT-012", "Derniere Maintenance", "Maintenance", "Derniere intervention ?", "Machine", "Date", "MAX(date_debut)", "Carte", "Fraicheur", "< 30 jours", "30-90 jours", "> 90 jours", "Oui", "maintenance", "date_debut", "ML-03"),
    ("MNT-013", "Frequence", "Maintenance", "Combien par mois ?", "Machine, Periode", "Interventions/mois", "COUNT/Nb_mois", "Barres", "Frequence", "-", "-", "-", "Non", "maintenance", "-", "-"),
    ("MNT-014", "Duree Moyenne", "Maintenance", "Duree moyenne intervention ?", "Machine, Periode", "Minutes", "AVG(duree)", "Carte", "Efficacite maintenance", "<= 60 min", "60-120 min", "> 120 min", "Oui", "maintenance", "duree", "-"),
    ("MNT-015", "Historique", "Maintenance", "Historique complet ?", "Machine, Periode", "Liste chronologique", "SELECT * FROM maintenance ORDER BY date", "Tableau + frise", "Historique", "-", "-", "-", "Non", "maintenance", "-", "-"),
]

SEN_KPIS = [
    ("SEN-001", "Temperature Actuelle", "Capteurs", "Temperature actuelle ?", "Machine", "Derniere lecture (C)", "DERNIERE(temperature)", "Carte + jauge", "Etat thermique", "< 60C", "60-80C", "> 80C", "Oui", "sensor_data", "temperature", "ML-04"),
    ("SEN-002", "Temperature Moyenne", "Capteurs", "Temperature moyenne ?", "Machine, Periode", "Moyenne (C)", "AVG(temperature)", "Carte", "Tendance thermique", "< 60C", "60-80C", "> 80C", "Oui", "sensor_data", "temperature", "ML-04"),
    ("SEN-003", "Temperature Max", "Capteurs", "Temperature maximale ?", "Machine, Periode", "Max (C)", "MAX(temperature)", "Carte", "Pic temperature", "< 70C", "70-80C", "> 80C", "Oui", "sensor_data", "temperature", "ML-04"),
    ("SEN-004", "Vibration Actuelle", "Capteurs", "Vibration actuelle ?", "Machine", "Derniere lecture (mm/s)", "DERNIERE(vibration)", "Carte + jauge", "Etat mecanique", "< 2.5", "2.5-4.5", "> 4.5", "Oui", "sensor_data", "vibration", "ML-04"),
    ("SEN-005", "Vibration Moyenne", "Capteurs", "Vibration moyenne ?", "Machine, Periode", "Moyenne (mm/s)", "AVG(vibration)", "Carte", "Tendance vibration", "< 2.5", "2.5-4.5", "> 4.5", "Oui", "sensor_data", "vibration", "ML-04"),
    ("SEN-006", "Vibration Max", "Capteurs", "Vibration maximale ?", "Machine, Periode", "Max (mm/s)", "MAX(vibration)", "Carte", "Pic vibration", "< 3.5", "3.5-4.5", "> 4.5", "Oui", "sensor_data", "vibration", "ML-04"),
    ("SEN-007", "RPM Moyen", "Capteurs", "Broche tourne correctement ?", "Machine, Periode", "RPM moyen", "AVG(rpm)", "Carte + courbe", "Vitesse broche", "Plage prevue", "Legere deviation", "Deviation majeure", "Oui", "sensor_data", "rpm", "-"),
    ("SEN-008", "Charge Broche", "Capteurs", "Broche surchargee ?", "Machine, Periode", "Charge %", "AVG(charge_frappe)", "Carte + courbe", "Charge mecanique", "< 80%", "80%-95%", "> 95%", "Oui", "sensor_data", "charge_frappe", "-"),
    ("SEN-009", "Puissance Moyenne", "Capteurs", "Consommation normale ?", "Machine, Periode", "Puissance kW", "AVG(puissance)", "Carte + courbe", "Consommation energetique", "Plage normale", "Legere augmentation", "Anormale", "Oui", "sensor_data", "puissance", "-"),
    ("SEN-010", "Temps Cycle Moyen", "Capteurs", "Temps cycle stable ?", "Machine, Periode", "Secondes", "AVG(temps_cycle)", "Carte + courbe", "Stabilite cycle", "Stable", "Variable", "Tres variable", "Oui", "sensor_data", "temps_cycle", "-"),
    ("SEN-011", "Score Anomalie", "Capteurs", "Anomalies detectees ?", "Machine", "Score 0-100", "Calcule sur capteurs", "Carte + jauge", "Sante globale", "< 30", "30-60", "> 60", "Oui", "sensor_data (calcule)", "-", "ML-04"),
    ("SEN-012", "Nb Alertes Capteurs", "Capteurs", "Depassements seuil ?", "Machine, Periode", "Nombre alertes", "COUNT(WHERE temp>vibration>seuil)", "Carte", "Anomalies", "0", "1-5", "> 5", "Oui", "sensor_data (calcule)", "-", "ML-04"),
    ("SEN-013", "Statut Machine (Capteurs)", "Capteurs", "Machine en marche ?", "Machine", "Statut", "DERNIERE(statut_machine)", "Badge colore", "Etat operationnel", "RUNNING", "STOPPED/MAINT", "BROKEN", "Non", "sensor_data", "statut_machine", "ML-04"),
    ("SEN-014", "Courbe Temperature", "Capteurs", "Evolution temperature ?", "Machine, Periode", "Serie temporelle", "temperature GROUP BY timestamp", "Courbe + seuils", "Tendance", "-", "-", "-", "Non", "sensor_data", "temperature", "ML-04"),
    ("SEN-015", "Courbe Vibration", "Capteurs", "Evolution vibration ?", "Machine, Periode", "Serie temporelle", "vibration GROUP BY timestamp", "Courbe + seuils", "Tendance mecanique", "-", "-", "-", "Non", "sensor_data", "vibration", "ML-04"),
]

ALL_MODULES = [
    ("Vue Executive", EXEC_KPIS),
    ("Machine", MCH_KPIS),
    ("Ordre Fabrication", OF_KPIS),
    ("Qualite", QUA_KPIS),
    ("Inventaire", INV_KPIS),
    ("Outil", TL_KPIS),
    ("Maintenance", MNT_KPIS),
    ("Capteurs", SEN_KPIS),
]

# ── Alert Data ──────────────────────────────────────────────────────────

ALERT_DATA = [
    ("A01", "Machine en panne", "Critique", "machine.statut = 'BROKEN'", "Appeler technicien maintenance"),
    ("A02", "Stock matiere critique", "Critique", "stock.quantite_stock <= seuil_alerte", "Commander reapprovisionnement"),
    ("A03", "Stock outil critique", "Critique", "stock_outil.quantite_stock <= seuil_alerte", "Commander nouvel outil"),
    ("A04", "Vibrations excessives", "Critique", "vibration > 4.5 mm/s", "Arreter machine, inspecter"),
    ("A05", "Temperature excessive", "Avertissement", "temperature > 80C", "Verifier refroidissement"),
    ("A06", "OEE bas", "Avertissement", "OEE < 60%", "Analyser les causes"),
    ("A07", "Taux de rebut eleve", "Avertissement", "taux_rebut > 5%", "Verifier qualite/outil/reglage"),
    ("A08", "Outil use", "Avertissement", "usure > 80%", "Planifier remplacement"),
    ("A09", "OF en retard", "Information", "date_fin_reelle > date_fin_prevue", "Reorganiser planning"),
    ("A10", "Maintenance en retard", "Information", "Date prevue depassee", "Planifier intervention"),
]

ALERT_COLUMNS = ["Alerte ID", "Description", "Severite", "Condition", "Action Recommandee"]

# ── ML Data ─────────────────────────────────────────────────────────────

ML_DATA = [
    ("ML-01", "Scrap Prediction", "Qualite", "Probabilite de rebut", "Parametres coupe, outil, capteurs, operateur"),
    ("ML-02", "Machining Time Estimation", "Machine/OF", "Temps d'usinage estime", "Gamme, machine, outil, matiere"),
    ("ML-03", "Predictive Maintenance", "Maintenance", "Temps avant prochaine panne", "Capteurs, historique, age machine"),
    ("ML-04", "Machine Failure Prediction", "Machine/Capteurs", "Probabilite panne (7 jours)", "Capteurs tendances, maintenance, execution"),
    ("ML-05", "Tool Wear Prediction", "Outil", "Usure restante predite", "Type outil, executions, parametres coupe"),
    ("ML-06", "Production Duration Prediction", "OF", "Duree totale predite", "Quantite, gamme, operateur, matiere"),
    ("ML-07", "Inventory Forecasting", "Inventaire", "Date de rupture stock", "Stock actuel, consommation, OFs actifs"),
]

ML_COLUMNS = ["Modele ID", "Nom du Modele", "Module Cible", "Prediction", "Variables d'Entree"]

# ── Summary Data ────────────────────────────────────────────────────────

SUMMARY_DATA = [
    ("Vue Executive", 15, "Vision globale de l'atelier", "OEE global, production, alertes"),
    ("Machine", 20, "Performance complete d'une machine", "Statut, OEE, capteurs, maintenance, outils"),
    ("Ordre Fabrication", 15, "Suivi d'un OF", "Quantites, delais, efficacite"),
    ("Qualite", 20, "Analyse qualite", "Rebut, conformite, FPY, causes"),
    ("Inventaire", 15, "Surveillance des stocks", "Stock, statut, valeur, reappro"),
    ("Outil", 15, "Etat des outils", "Usure, duree de vie, executions"),
    ("Maintenance", 15, "Gestion maintenance", "Historique, MTBF, MTTR, couts"),
    ("Capteurs", 15, "Donnees temps reel", "Temperature, vibration, RPM, anomalies"),
]

SUMMARY_COLUMNS = ["Module", "Nombre KPI", "Objectif", "Contenu Principal"]


# ── Excel generation ────────────────────────────────────────────────────

def apply_header_style(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def apply_body_style(ws, start_row, end_row, num_cols):
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER


def color_seuil_cells(ws, row, green_col, orange_col, red_col):
    ws.cell(row=row, column=green_col).fill = GREEN_FILL
    ws.cell(row=row, column=orange_col).fill = ORANGE_FILL
    ws.cell(row=row, column=red_col).fill = RED_FILL


def set_column_widths(ws, num_cols, widths=None):
    for col in range(1, num_cols + 1):
        if widths and col <= len(widths):
            ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]
        else:
            ws.column_dimensions[get_column_letter(col)].width = 18


def create_summary_sheet(wb):
    ws = wb.active
    ws.title = "Resume"
    ws.merge_cells("A1:D1")
    ws["A1"] = "AMIP - Catalogue KPI Dashboard"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Version 1.0 | 2026-07-15 | 130 KPIs | 8 Modules"
    ws["A2"].font = SUBTITLE_FONT
    ws["A4"] = "Modules"
    ws["A4"].font = SUBTITLE_FONT
    for col_idx, col_name in enumerate(SUMMARY_COLUMNS, 1):
        ws.cell(row=5, column=col_idx, value=col_name)
    apply_header_style(ws, 5, len(SUMMARY_COLUMNS))
    for row_idx, row_data in enumerate(SUMMARY_DATA, 6):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    apply_body_style(ws, 6, 5 + len(SUMMARY_DATA), len(SUMMARY_COLUMNS))
    total_row = 6 + len(SUMMARY_DATA)
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=2, value=sum(r[1] for r in SUMMARY_DATA))
    ws.cell(row=total_row, column=1).font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=total_row, column=2).font = Font(name="Calibri", bold=True, size=11)
    set_column_widths(ws, len(SUMMARY_COLUMNS), [25, 15, 40, 45])


def create_kpi_sheet(wb, module_name, kpis):
    ws = wb.create_sheet(title=module_name[:31])
    ws.merge_cells(f"A1:{get_column_letter(len(KPI_COLUMNS))}1")
    ws["A1"] = f"Module: {module_name}"
    ws["A1"].font = TITLE_FONT
    for col_idx, col_name in enumerate(KPI_COLUMNS, 1):
        ws.cell(row=3, column=col_idx, value=col_name)
    apply_header_style(ws, 3, len(KPI_COLUMNS))
    for row_idx, kpi in enumerate(kpis, 4):
        for col_idx, val in enumerate(kpi, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    apply_body_style(ws, 4, 3 + len(kpis), len(KPI_COLUMNS))
    for row_idx in range(4, 4 + len(kpis)):
        color_seuil_cells(ws, row_idx, 10, 11, 12)
    set_column_widths(ws, len(KPI_COLUMNS), [12, 28, 18, 40, 22, 35, 35, 18, 30, 15, 15, 15, 12, 25, 30, 10])


def create_alert_sheet(wb):
    ws = wb.create_sheet(title="Alertes")
    ws.merge_cells("A1:E1")
    ws["A1"] = "Systeme d'Alertes"
    ws["A1"].font = TITLE_FONT
    for col_idx, col_name in enumerate(ALERT_COLUMNS, 1):
        ws.cell(row=3, column=col_idx, value=col_name)
    apply_header_style(ws, 3, len(ALERT_COLUMNS))
    sever_colors = {"Critique": RED_FILL, "Avertissement": ORANGE_FILL, "Information": GREEN_FILL}
    for row_idx, alert in enumerate(ALERT_DATA, 4):
        for col_idx, val in enumerate(alert, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    apply_body_style(ws, 4, 3 + len(ALERT_DATA), len(ALERT_COLUMNS))
    for row_idx in range(4, 4 + len(ALERT_DATA)):
        sev = ws.cell(row=row_idx, column=3).value
        if sev in sever_colors:
            ws.cell(row=row_idx, column=3).fill = sever_colors[sev]
    set_column_widths(ws, len(ALERT_COLUMNS), [12, 28, 15, 40, 40])


def create_ml_sheet(wb):
    ws = wb.create_sheet(title="ML Predictions")
    ws.merge_cells("A1:E1")
    ws["A1"] = "Modeles Machine Learning"
    ws["A1"].font = TITLE_FONT
    for col_idx, col_name in enumerate(ML_COLUMNS, 1):
        ws.cell(row=3, column=col_idx, value=col_name)
    apply_header_style(ws, 3, len(ML_COLUMNS))
    for row_idx, ml in enumerate(ML_DATA, 4):
        for col_idx, val in enumerate(ml, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    apply_body_style(ws, 4, 3 + len(ML_DATA), len(ML_COLUMNS))
    set_column_widths(ws, len(ML_COLUMNS), [12, 30, 20, 35, 45])


def create_thresholds_sheet(wb):
    ws = wb.create_sheet(title="Seuils Configurables")
    ws.merge_cells("A1:G1")
    ws["A1"] = "Seuils Configurables"
    ws["A1"].font = TITLE_FONT
    cols = ["KPI ID", "Nom KPI", "Seuil Vert", "Seuil Orange", "Seuil Rouge", "Configurable", "Valeur Defaut Industrie"]
    for col_idx, col_name in enumerate(cols, 1):
        ws.cell(row=3, column=col_idx, value=col_name)
    apply_header_style(ws, 3, len(cols))
    threshold_kpis = []
    for mod_name, kpis in ALL_MODULES:
        for kpi in kpis:
            if kpi[12] == "Oui":
                threshold_kpis.append(kpi)
    for row_idx, kpi in enumerate(threshold_kpis, 4):
        ws.cell(row=row_idx, column=1, value=kpi[0])
        ws.cell(row=row_idx, column=2, value=kpi[1])
        ws.cell(row=row_idx, column=3, value=kpi[9])
        ws.cell(row=row_idx, column=4, value=kpi[10])
        ws.cell(row=row_idx, column=5, value=kpi[11])
        ws.cell(row=row_idx, column=6, value="Oui")
        ws.cell(row=row_idx, column=7, value=kpi[9])
    apply_body_style(ws, 4, 3 + len(threshold_kpis), len(cols))
    for row_idx in range(4, 4 + len(threshold_kpis)):
        ws.cell(row=row_idx, column=3).fill = GREEN_FILL
        ws.cell(row=row_idx, column=4).fill = ORANGE_FILL
        ws.cell(row=row_idx, column=5).fill = RED_FILL
    set_column_widths(ws, len(cols), [12, 30, 15, 15, 15, 15, 25])


def main():
    output = Path(__file__).parent / "kpi_catalog.xlsx"
    wb = openpyxl.Workbook()
    create_summary_sheet(wb)
    for mod_name, kpis in ALL_MODULES:
        create_kpi_sheet(wb, mod_name, kpis)
    create_alert_sheet(wb)
    create_ml_sheet(wb)
    create_thresholds_sheet(wb)
    wb.save(output)
    total_kpis = sum(len(kpis) for _, kpis in ALL_MODULES)
    print(f"Output: {output} ({output.stat().st_size // 1024} KB)")
    print(f"KPIs: {total_kpis}")
    print(f"Sheets: {len(wb.sheetnames)}")
    print("Sheets:", ", ".join(wb.sheetnames))


if __name__ == "__main__":
    main()
