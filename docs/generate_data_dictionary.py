import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from config.settings import DOCS_DIR, GENERATED_DATA_DIR
import csv

TABLES = {
    "secteur": [
        ("secteur_id", "SERIAL", "PK", "", "Identifiant secteur"),
        ("code", "VARCHAR(10)", "UNIQUE, NOT NULL", "", "Code court du secteur"),
        ("nom", "VARCHAR(100)", "NOT NULL", "", "Nom du secteur"),
        ("description", "TEXT", "", "", "Description"),
        ("created_at", "TIMESTAMP", "DEFAULT NOW()", "", "Date de creation"),
        ("updated_at", "TIMESTAMP", "DEFAULT NOW()", "", "Derniere mise a jour"),
    ],
    "machine": [
        ("machine_id", "SERIAL", "PK", "", "Identifiant machine"),
        ("code", "VARCHAR(20)", "UNIQUE, NOT NULL", "", "Code machine (M001, M002...)"),
        ("nom", "VARCHAR(100)", "NOT NULL", "", "Nom de la machine"),
        ("type", "VARCHAR(50)", "NOT NULL", "", "Tour CNC / Centre usinage CNC / Fraiseuse CNC"),
        ("marque", "VARCHAR(50)", "NOT NULL", "", "Marque (HANQI-CNC, Hartford)"),
        ("modele", "VARCHAR(50)", "NOT NULL", "", "Modele"),
        ("controller", "VARCHAR(50)", "", "", "Controleur CNC"),
        ("axes", "INTEGER", "", "", "Nombre d'axes"),
        ("rpm_max", "INTEGER", "", "", "Regime max broche (tr/min)"),
        ("tool_capacity", "INTEGER", "DEFAULT 0", "", "Capacite magazine outils"),
        ("statut", "VARCHAR(20)", "NOT NULL", "", "RUNNING / STOPPED / MAINTENANCE / BROKEN"),
        ("secteur_id", "INTEGER", "FK", "secteur", "Secteur d'affectation"),
        ("date_installation", "DATE", "", "", "Date d'installation"),
        ("created_at", "TIMESTAMP", "DEFAULT NOW()", "", "Date de creation"),
        ("updated_at", "TIMESTAMP", "DEFAULT NOW()", "", "Derniere mise a jour"),
    ],
    "operateur": [
        ("operateur_id", "SERIAL", "PK", "", "Identifiant operateur"),
        ("matricule", "VARCHAR(20)", "UNIQUE, NOT NULL", "", "Matricule employe"),
        ("nom", "VARCHAR(100)", "NOT NULL", "", "Nom"),
        ("prenom", "VARCHAR(100)", "NOT NULL", "", "Prenom"),
        ("poste", "VARCHAR(50)", "", "", "Operateur CNC / Controleur / Maintenance"),
        ("niveau_competence", "VARCHAR(20)", "", "", "Junior / Confirme / Senior"),
        ("date_embauche", "DATE", "", "", "Date d'embauche"),
        ("actif", "BOOLEAN", "DEFAULT TRUE", "", "Employe actif"),
        ("created_at", "TIMESTAMP", "DEFAULT NOW()", "", "Date de creation"),
        ("updated_at", "TIMESTAMP", "DEFAULT NOW()", "", "Derniere mise a jour"),
    ],
    "matiere": [
        ("matiere_id", "SERIAL", "PK", "", "Identifiant matiere"),
        ("code", "VARCHAR(20)", "UNIQUE, NOT NULL", "", "Code matiere"),
        ("designation", "VARCHAR(100)", "NOT NULL", "", "Description"),
        ("type_matiere", "VARCHAR(50)", "", "", "Acier / Aluminium / Inox / Cuivre / Plastique"),
        ("nuance", "VARCHAR(50)", "", "", "Specification nuance"),
        ("densite", "DECIMAL(6,3)", "", "", "Densite g/cm3"),
        ("prix_kg", "DECIMAL(10,2)", "", "", "Prix au kg"),
        ("unite", "VARCHAR(10)", "DEFAULT 'kg'", "", "Unite de mesure"),
        ("created_at", "TIMESTAMP", "DEFAULT NOW()", "", "Date de creation"),
        ("updated_at", "TIMESTAMP", "DEFAULT NOW()", "", "Derniere mise a jour"),
    ],
    "outil": [
        ("outil_id", "SERIAL", "PK", "", "Identifiant outil"),
        ("code", "VARCHAR(20)", "UNIQUE, NOT NULL", "", "Code outil"),
        ("designation", "VARCHAR(100)", "NOT NULL", "", "Description outil"),
        ("type_outil", "VARCHAR(50)", "NOT NULL", "", "Foret / Fraise / Alesoir / Taraud / Plateau / Mandrin"),
        ("diametre", "DECIMAL(8,3)", "", "", "Diametre en mm"),
        ("matiere_outil", "VARCHAR(50)", "", "", "HSS / Carbure / Ceramique / Diamant"),
        ("duree_vie_totale", "INTEGER", "NOT NULL", "", "Duree de vie totale (minutes)"),
        ("usure_actuelle", "INTEGER", "DEFAULT 0", "", "Usure actuelle (minutes)"),
        ("duree_vie_restante", "INTEGER", "NOT NULL", "", "Duree de vie restante (minutes)"),
        ("cout_achat", "DECIMAL(10,2)", "", "", "Cout d'achat"),
        ("cout_remplacement", "DECIMAL(10,2)", "", "", "Cout de remplacement"),
        ("disponible", "BOOLEAN", "DEFAULT TRUE", "", "Disponible pour utilisation"),
        ("created_at", "TIMESTAMP", "DEFAULT NOW()", "", "Date de creation"),
        ("updated_at", "TIMESTAMP", "DEFAULT NOW()", "", "Derniere mise a jour"),
    ],
    "stock_outil": [
        ("stock_outil_id", "SERIAL", "PK", "", "Identifiant stock outil"),
        ("outil_id", "INTEGER", "FK, UNIQUE", "outil", "Outil reference"),
        ("quantite_stock", "INTEGER", "NOT NULL, DEFAULT 0", "", "Quantite en stock"),
        ("emplacement", "VARCHAR(50)", "", "", "Emplacement de stockage"),
        ("seuil_alerte", "INTEGER", "DEFAULT 5", "", "Seuil de reapprovisionnement"),
        ("date_derniere_maj", "TIMESTAMP", "", "", "Derniere mise a jour stock"),
        ("created_at", "TIMESTAMP", "DEFAULT NOW()", "", "Date de creation"),
        ("updated_at", "TIMESTAMP", "DEFAULT NOW()", "", "Derniere mise a jour"),
    ],
    "piece": [
        ("piece_id", "SERIAL", "PK", "", "Identifiant piece"),
        ("reference", "VARCHAR(30)", "UNIQUE, NOT NULL", "", "Reference piece"),
        ("designation", "VARCHAR(150)", "NOT NULL", "", "Description piece"),
        ("famille", "VARCHAR(50)", "", "", "Famille de pieces"),
        ("matiere_id", "INTEGER", "FK", "matiere", "Matiere premiere"),
        ("poids", "DECIMAL(10,4)", "", "", "Poids en kg"),
        ("dimensions", "VARCHAR(100)", "", "", "Dimensions Lxlxh mm"),
        ("tolerances", "VARCHAR(100)", "", "", "Tolerances"),
        ("prix_revient", "DECIMAL(10,2)", "", "", "Prix de revient"),
        ("created_at", "TIMESTAMP", "DEFAULT NOW()", "", "Date de creation"),
        ("updated_at", "TIMESTAMP", "DEFAULT NOW()", "", "Derniere mise a jour"),
    ],
    "programme_usinage": [
        ("programme_id", "SERIAL", "PK", "", "Identifiant programme"),
        ("code_programme", "VARCHAR(30)", "UNIQUE, NOT NULL", "", "Code programme"),
        ("nom", "VARCHAR(100)", "NOT NULL", "", "Nom du programme"),
        ("version", "VARCHAR(10)", "", "", "Numero de version"),
        ("description", "TEXT", "", "", "Description"),
        ("duree_estimee", "INTEGER", "", "", "Duree estimee (minutes)"),
    ],
    "gamme_usinage": [
        ("gamme_id", "SERIAL", "PK", "", "Identifiant gamme"),
        ("code", "VARCHAR(20)", "UNIQUE, NOT NULL", "", "Code gamme"),
        ("designation", "VARCHAR(100)", "NOT NULL", "", "Description gamme"),
        ("piece_id", "INTEGER", "FK, NOT NULL", "piece", "Piece concernee"),
        ("nb_phases", "INTEGER", "", "", "Nombre de phases"),
        ("duree_totale_estimee", "INTEGER", "", "", "Duree totale estimee (min)"),
        ("version", "VARCHAR(10)", "DEFAULT '1.0'", "", "Version"),
        ("statut", "VARCHAR(20)", "DEFAULT 'ACTIVE'", "", "ACTIVE / OBSOLETE"),
        ("created_at", "TIMESTAMP", "DEFAULT NOW()", "", "Date de creation"),
        ("updated_at", "TIMESTAMP", "DEFAULT NOW()", "", "Derniere mise a jour"),
    ],
    "phase_gamme": [
        ("phase_gamme_id", "SERIAL", "PK", "", "Identifiant phase"),
        ("gamme_id", "INTEGER", "FK, NOT NULL", "gamme_usinage", "Gamme parente"),
        ("numero_phase", "INTEGER", "NOT NULL", "", "Numero sequence phase"),
        ("designation", "VARCHAR(100)", "", "", "Description phase"),
        ("machine_id", "INTEGER", "FK, NOT NULL", "machine", "Machine planifiee"),
        ("outil_id", "INTEGER", "FK, NOT NULL", "outil", "Outil planifie"),
        ("programme_id", "INTEGER", "FK", "programme_usinage", "Programme CNC"),
        ("temps_usinage_prevu", "INTEGER", "NOT NULL", "", "Temps usinage prevu (min)"),
        ("temps_reglage_prevu", "INTEGER", "NOT NULL", "", "Temps reglage prevu (min)"),
        ("exigence_technique", "TEXT", "", "", "Exigences techniques"),
        ("created_at", "TIMESTAMP", "DEFAULT NOW()", "", "Date de creation"),
        ("updated_at", "TIMESTAMP", "DEFAULT NOW()", "", "Derniere mise a jour"),
    ],
    "ordre_fabrication": [
        ("ordre_fabrication_id", "SERIAL", "PK", "", "Identifiant OF"),
        ("numero_of", "VARCHAR(20)", "UNIQUE, NOT NULL", "", "Numero ordre fabrication"),
        ("piece_id", "INTEGER", "FK, NOT NULL", "piece", "Piece a fabriquer"),
        ("gamme_id", "INTEGER", "FK, NOT NULL", "gamme_usinage", "Gamme a suivre"),
        ("quantite_demandee", "INTEGER", "NOT NULL", "", "Quantite demandee"),
        ("quantite_produite", "INTEGER", "DEFAULT 0", "", "Quantite produite"),
        ("quantite_rebut", "INTEGER", "DEFAULT 0", "", "Quantite rebut"),
        ("date_debut_prevue", "DATE", "NOT NULL", "", "Date debut prevue"),
        ("date_fin_prevue", "DATE", "", "", "Date fin prevue"),
        ("date_debut_reelle", "DATE", "", "", "Date debut reelle"),
        ("date_fin_reelle", "DATE", "", "", "Date fin reelle"),
        ("priorite", "VARCHAR(20)", "DEFAULT 'NORMALE'", "", "HAUTE / NORMALE / BASSE"),
        ("statut", "VARCHAR(20)", "NOT NULL", "", "EN_ATTENTE / EN_COURS / TERMINE / ANNULE"),
        ("created_at", "TIMESTAMP", "DEFAULT NOW()", "", "Date de creation"),
        ("updated_at", "TIMESTAMP", "DEFAULT NOW()", "", "Derniere mise a jour"),
    ],
    "execution_phase": [
        ("execution_id", "SERIAL", "PK", "", "Identifiant execution"),
        ("ordre_fabrication_id", "INTEGER", "FK, NOT NULL", "ordre_fabrication", "OF reference"),
        ("phase_gamme_id", "INTEGER", "FK, NOT NULL", "phase_gamme", "Phase executee"),
        ("machine_id", "INTEGER", "FK, NOT NULL", "machine", "Machine reelle"),
        ("outil_id", "INTEGER", "FK", "outil", "Outil reel"),
        ("operateur_id", "INTEGER", "FK", "operateur", "Operateur"),
        ("programme_id", "INTEGER", "FK", "programme_usinage", "Programme utilise"),
        ("date_debut", "TIMESTAMP", "NOT NULL", "", "Date debut reelle"),
        ("date_fin", "TIMESTAMP", "", "", "Date fin reelle"),
        ("temps_usinage_reel", "INTEGER", "", "", "Temps usinage reel (min)"),
        ("temps_reglage_reel", "INTEGER", "", "", "Temps reglage reel (min)"),
        ("nb_pieces_produites", "INTEGER", "DEFAULT 0", "", "Pieces produites"),
        ("nb_pieces_rebut", "INTEGER", "DEFAULT 0", "", "Pieces rebut"),
        ("vitesse_coupe", "DECIMAL(10,2)", "", "", "Vitesse de coupe (m/min)"),
        ("avance", "DECIMAL(10,3)", "", "", "Avance (mm/tr)"),
        ("profondeur_passe", "DECIMAL(8,3)", "", "", "Profondeur de passe (mm)"),
        ("statut", "VARCHAR(20)", "", "", "EN_COURS / TERMINE / ARRET"),
        ("created_at", "TIMESTAMP", "DEFAULT NOW()", "", "Date de creation"),
        ("updated_at", "TIMESTAMP", "DEFAULT NOW()", "", "Derniere mise a jour"),
    ],
    "execution_outil": [
        ("execution_outil_id", "SERIAL", "PK", "", "Identifiant execution outil"),
        ("execution_id", "INTEGER", "FK, NOT NULL", "execution_phase", "Execution reference"),
        ("outil_id", "INTEGER", "FK, NOT NULL", "outil", "Outil utilise"),
        ("usure_debut", "INTEGER", "", "", "Usure au debut (min)"),
        ("usure_fin", "INTEGER", "", "", "Usure a la fin (min)"),
        ("duree_utilisation", "INTEGER", "", "", "Duree utilisation (min)"),
        ("created_at", "TIMESTAMP", "DEFAULT NOW()", "", "Date de creation"),
    ],
    "cause_rebut": [
        ("cause_rebut_id", "SERIAL", "PK", "", "Identifiant cause"),
        ("code", "VARCHAR(20)", "UNIQUE, NOT NULL", "", "Code cause"),
        ("categorie", "VARCHAR(50)", "NOT NULL", "", "Materiel / Outil / Machine / Programmation / Operateur / Autre"),
        ("description", "VARCHAR(200)", "NOT NULL", "", "Description cause"),
        ("created_at", "TIMESTAMP", "DEFAULT NOW()", "", "Date de creation"),
    ],
    "controle_qualite": [
        ("controle_id", "SERIAL", "PK", "", "Identifiant controle"),
        ("execution_id", "INTEGER", "FK, NOT NULL", "execution_phase", "Execution reference"),
        ("piece_id", "INTEGER", "FK, NOT NULL", "piece", "Piece controlee"),
        ("cause_rebut_id", "INTEGER", "FK", "cause_rebut", "Cause defeaut (si non conforme)"),
        ("date_controle", "TIMESTAMP", "NOT NULL", "", "Date du controle"),
        ("resultat", "VARCHAR(20)", "NOT NULL", "", "CONFORME / NON_CONFORME / EN_ATTENTE"),
        ("nb_controles", "INTEGER", "DEFAULT 0", "", "Nombre pieces controlees"),
        ("nb_conformes", "INTEGER", "DEFAULT 0", "", "Nombre conformes"),
        ("nb_non_conformes", "INTEGER", "DEFAULT 0", "", "Nombre non conformes"),
        ("dimension_mesuree", "DECIMAL(10,3)", "", "", "Dimension mesuree"),
        ("dimension_cible", "DECIMAL(10,3)", "", "", "Dimension cible"),
        ("tolerance_plus", "DECIMAL(10,3)", "", "", "Tolerance +"),
        ("tolerance_moins", "DECIMAL(10,3)", "", "", "Tolerance -"),
        ("rugosite_mesuree", "DECIMAL(6,3)", "", "", "Rugosite Ra mesuree"),
        ("commentaire", "TEXT", "", "", "Commentaires"),
        ("created_at", "TIMESTAMP", "DEFAULT NOW()", "", "Date de creation"),
    ],
    "maintenance": [
        ("maintenance_id", "SERIAL", "PK", "", "Identifiant maintenance"),
        ("machine_id", "INTEGER", "FK, NOT NULL", "machine", "Machine concernee"),
        ("type_maintenance", "VARCHAR(30)", "NOT NULL", "", "Preventive / Corrective / Changement huile / ..."),
        ("description", "TEXT", "NOT NULL", "", "Description intervention"),
        ("date_debut", "TIMESTAMP", "NOT NULL", "", "Date debut"),
        ("date_fin", "TIMESTAMP", "", "", "Date fin"),
        ("duree", "INTEGER", "", "", "Duree (minutes)"),
        ("cout", "DECIMAL(10,2)", "", "", "Cout intervention"),
        ("operateur_id", "INTEGER", "FK", "operateur", "Technicien assigne"),
        ("statut", "VARCHAR(20)", "DEFAULT 'PLANIFIEE'", "", "PLANIFIEE / EN_COURS / TERMINEE"),
        ("cree_par", "VARCHAR(50)", "", "", "Demandeur"),
        ("created_at", "TIMESTAMP", "DEFAULT NOW()", "", "Date de creation"),
        ("updated_at", "TIMESTAMP", "DEFAULT NOW()", "", "Derniere mise a jour"),
    ],
    "sensor_data": [
        ("sensor_id", "BIGSERIAL", "PK", "", "Identifiant capteur"),
        ("machine_id", "INTEGER", "FK, NOT NULL", "machine", "Machine source"),
        ("timestamp", "TIMESTAMP", "NOT NULL", "", "Horodatage mesure"),
        ("temperature", "DECIMAL(6,2)", "", "", "Temperature broche (C)"),
        ("vibration", "DECIMAL(6,3)", "", "", "Vibration (mm/s)"),
        ("rpm", "INTEGER", "", "", "Regime broche (tr/min)"),
        ("charge_frappe", "DECIMAL(6,2)", "", "", "Charge broche (%)"),
        ("puissance", "DECIMAL(8,2)", "", "", "Consommation electrique (kW)"),
        ("vitesse_avance", "DECIMAL(8,2)", "", "", "Vitesse avance (mm/min)"),
        ("statut_machine", "VARCHAR(20)", "", "", "RUNNING / STOPPED / MAINTENANCE / BROKEN"),
        ("temps_cycle", "DECIMAL(8,2)", "", "", "Temps cycle (secondes)"),
    ],
    "stock_piece": [
        ("stock_piece_id", "SERIAL", "PK", "", "Identifiant stock piece"),
        ("piece_id", "INTEGER", "FK, UNIQUE", "piece", "Piece reference"),
        ("quantite_stock", "INTEGER", "NOT NULL, DEFAULT 0", "", "Quantite en stock"),
        ("emplacement", "VARCHAR(50)", "", "", "Emplacement de stockage"),
        ("date_derniere_maj", "TIMESTAMP", "", "", "Derniere mise a jour stock"),
        ("created_at", "TIMESTAMP", "DEFAULT NOW()", "", "Date de creation"),
        ("updated_at", "TIMESTAMP", "DEFAULT NOW()", "", "Derniere mise a jour"),
    ],
    "stock_matiere": [
        ("stock_matiere_id", "SERIAL", "PK", "", "Identifiant stock matiere"),
        ("matiere_id", "INTEGER", "FK, UNIQUE", "matiere", "Matiere reference"),
        ("quantite_stock", "DECIMAL(12,3)", "NOT NULL, DEFAULT 0", "", "Stock en kg"),
        ("emplacement", "VARCHAR(50)", "", "", "Emplacement de stockage"),
        ("seuil_alerte", "DECIMAL(12,3)", "", "", "Stock minimum"),
        ("date_derniere_maj", "TIMESTAMP", "", "", "Derniere mise a jour stock"),
        ("created_at", "TIMESTAMP", "DEFAULT NOW()", "", "Date de creation"),
        ("updated_at", "TIMESTAMP", "DEFAULT NOW()", "", "Derniere mise a jour"),
    ],
}


def create_data_dictionary():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Dictionary"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    table_font = Font(bold=True, size=12, color="2F5496")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Table", "Colonne", "Type", "Contraintes", "FK Reference", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row = 2
    for table_name, columns in TABLES.items():
        for col_data in columns:
            ws.cell(row=row, column=1, value=table_name).border = thin_border
            for col_idx, val in enumerate(col_data, 2):
                ws.cell(row=row, column=col_idx, value=val).border = thin_border
            row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 45

    ws2 = wb.create_sheet("Table Summary")
    ws2.cell(row=1, column=1, value="Table").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Colonnes").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=3, value="CSV Rows").font = header_font
    ws2.cell(row=1, column=3).fill = header_fill
    ws2.cell(row=1, column=4, value="Description").font = header_font
    ws2.cell(row=1, column=4).fill = header_fill

    descriptions = {
        "secteur": "Secteurs de l'atelier",
        "machine": "Machines CNC",
        "operateur": "Operateurs et personnel",
        "matiere": "Matieres premieres",
        "outil": "Outillage de usinage",
        "stock_outil": "Stock outillage",
        "piece": "Pieces finies",
        "programme_usinage": "Programmes CNC",
        "gamme_usinage": "Gammes d'usinage",
        "phase_gamme": "Phases de gamme",
        "ordre_fabrication": "Ordres de fabrication",
        "execution_phase": "Executions phases",
        "execution_outil": "Consommation outils",
        "cause_rebut": "Causes de rebut",
        "controle_qualite": "Controles qualite",
        "maintenance": "Maintenances",
        "sensor_data": "Donnees capteurs IoT",
        "stock_piece": "Stock pieces finies",
        "stock_matiere": "Stock matieres",
    }

    for r, (table_name, columns) in enumerate(TABLES.items(), 2):
        csv_file = GENERATED_DATA_DIR / f"{table_name}.csv"
        csv_rows = 0
        if csv_file.exists():
            with open(csv_file, "r", encoding="utf-8") as f:
                csv_rows = sum(1 for _ in f) - 1

        ws2.cell(row=r, column=1, value=table_name).border = thin_border
        ws2.cell(row=r, column=2, value=len(columns)).border = thin_border
        ws2.cell(row=r, column=3, value=csv_rows).border = thin_border
        ws2.cell(row=r, column=4, value=descriptions.get(table_name, "")).border = thin_border

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 40

    output_path = DOCS_DIR / "data_dictionary.xlsx"
    wb.save(output_path)
    print(f"Data dictionary created: {output_path}")


if __name__ == "__main__":
    create_data_dictionary()
