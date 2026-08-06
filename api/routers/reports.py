import io
from datetime import date
from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from api.database import fetch_all, fetch_one

router = APIRouter(prefix="/api/reports", tags=["Export Rapports"])

# ─── helpers ────────────────────────────────────────────────────────────

def _excel_header(ws, headers):
    from openpyxl.styles import Font, PatternFill
    fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    font = Font(color="FFFFFF", bold=True, size=11)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = fill
        cell.font = font

def _make_excel(title, headers, rows):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    _excel_header(ws, headers)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def _make_pdf(title, headers, rows):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=10*mm, leftMargin=10*mm, topMargin=15*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    elements = [Paragraph(f"<b>{title}</b>", styles["Title"]), Spacer(1, 8*mm)]

    data = [headers] + [[str(c or "") for c in r] for r in rows]
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F9")]),
    ]))
    elements.append(tbl)
    doc.build(elements)
    buf.seek(0)
    return buf

# ─── Quality Reports ────────────────────────────────────────────────────

@router.get("/quality/excel")
def quality_excel():
    rows = fetch_all("""
        SELECT c.date_controle::text AS dt, ma.code, p.reference, p.designation,
               c.resultat, (c.dimension_mesuree - c.dimension_cible) AS ecart, c.rugosite_mesuree
        FROM controle_qualite c
        JOIN execution_phase e ON c.execution_id = e.execution_id
        JOIN machine ma ON e.machine_id = ma.machine_id
        JOIN piece p ON c.piece_id = p.piece_id
        ORDER BY c.date_controle DESC LIMIT 5000
    """)
    headers = ["Date", "Machine", "Reference", "Piece", "Resultat", "Ecart (mm)", "Rugosite"]
    data = [[r.get("dt",""), r.get("code",""), r.get("reference",""), r.get("designation",""),
             r.get("resultat",""), r.get("ecart",""), r.get("rugosite_mesuree","")] for r in rows]
    buf = _make_excel("Qualite", headers, data)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=rapport_qualite.xlsx"})

@router.get("/quality/pdf")
def quality_pdf():
    rows = fetch_all("""
        SELECT c.date_controle::text AS dt, ma.code, p.reference, p.designation,
               c.resultat, (c.dimension_mesuree - c.dimension_cible) AS ecart, c.rugosite_mesuree
        FROM controle_qualite c
        JOIN execution_phase e ON c.execution_id = e.execution_id
        JOIN machine ma ON e.machine_id = ma.machine_id
        JOIN piece p ON c.piece_id = p.piece_id
        ORDER BY c.date_controle DESC LIMIT 2000
    """)
    headers = ["Date", "Machine", "Reference", "Piece", "Resultat", "Ecart (mm)", "Rugosite"]
    data = [[r.get("dt",""), r.get("code",""), r.get("reference",""), r.get("designation",""),
             r.get("resultat",""), r.get("ecart",""), r.get("rugosite_mesuree","")] for r in rows]
    buf = _make_pdf("Rapport Qualite", headers, data)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=rapport_qualite.pdf"})

# ─── Maintenance Reports ────────────────────────────────────────────────

@router.get("/maintenance/excel")
def maintenance_excel():
    rows = fetch_all("""
        SELECT m.date_debut::text AS dt, m.type_maintenance, ma.code,
               m.description, m.cout, m.duree,
               CONCAT(o.prenom, ' ', o.nom) AS responsable
        FROM maintenance m
        JOIN machine ma ON m.machine_id = ma.machine_id
        LEFT JOIN operateur o ON m.operateur_id = o.operateur_id
        ORDER BY m.date_debut DESC LIMIT 5000
    """)
    headers = ["Date", "Type", "Machine", "Description", "Cout (€)", "Duree (h)", "Responsable"]
    data = [[r.get("dt",""), r.get("type_maintenance",""), r.get("code",""),
             r.get("description",""), r.get("cout",""), r.get("duree",""), r.get("responsable","")] for r in rows]
    buf = _make_excel("Maintenance", headers, data)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=rapport_maintenance.xlsx"})

@router.get("/maintenance/pdf")
def maintenance_pdf():
    rows = fetch_all("""
        SELECT m.date_debut::text AS dt, m.type_maintenance, ma.code,
               m.description, m.cout, m.duree,
               CONCAT(o.prenom, ' ', o.nom) AS responsable
        FROM maintenance m
        JOIN machine ma ON m.machine_id = ma.machine_id
        LEFT JOIN operateur o ON m.operateur_id = o.operateur_id
        ORDER BY m.date_debut DESC LIMIT 2000
    """)
    headers = ["Date", "Type", "Machine", "Description", "Cout (€)", "Duree (h)", "Responsable"]
    data = [[r.get("dt",""), r.get("type_maintenance",""), r.get("code",""),
             r.get("description",""), r.get("cout",""), r.get("duree",""), r.get("responsable","")] for r in rows]
    buf = _make_pdf("Rapport Maintenance", headers, data)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=rapport_maintenance.pdf"})

# ─── Production Reports ─────────────────────────────────────────────────

@router.get("/production/excel")
def production_excel():
    rows = fetch_all("""
        SELECT o.numero_of, p.reference, p.designation, o.quantite_demandee,
               o.date_debut_prevue::text, o.date_fin_prevue::text,
               o.statut, o.quantite_produite
        FROM ordre_fabrication o
        JOIN piece p ON o.piece_id = p.piece_id
        ORDER BY o.date_debut_prevue DESC LIMIT 5000
    """)
    headers = ["OF", "Reference", "Piece", "Qte demandee", "Date debut prevue", "Date fin prevue", "Statut", "Qte produite"]
    data = [[r.get("numero_of",""), r.get("reference",""), r.get("designation",""), r.get("quantite_demandee",""),
             r.get("date_debut_prevue",""), r.get("date_fin_prevue",""), r.get("statut",""), r.get("quantite_produite","")] for r in rows]
    buf = _make_excel("Production", headers, data)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=rapport_production.xlsx"})

@router.get("/production/pdf")
def production_pdf():
    rows = fetch_all("""
        SELECT o.numero_of, p.reference, p.designation, o.quantite_demandee,
               o.date_debut_prevue::text, o.date_fin_prevue::text,
               o.statut, o.quantite_produite
        FROM ordre_fabrication o
        JOIN piece p ON o.piece_id = p.piece_id
        ORDER BY o.date_debut_prevue DESC LIMIT 2000
    """)
    headers = ["OF", "Reference", "Piece", "Qte demandee", "Date debut prevue", "Date fin prevue", "Statut", "Qte produite"]
    data = [[r.get("numero_of",""), r.get("reference",""), r.get("designation",""), r.get("quantite_demandee",""),
             r.get("date_debut_prevue",""), r.get("date_fin_prevue",""), r.get("statut",""), r.get("quantite_produite","")] for r in rows]
    buf = _make_pdf("Rapport Production", headers, data)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=rapport_production.pdf"})
